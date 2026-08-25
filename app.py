import os
import re
import math
import base64
import io
import zipfile
import tempfile
import contextlib
from pathlib import Path

import gradio as gr
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.utils.protection import hash_password

PASSWORD = "GLBA"
DATA_LAST_ROW = 5000
VBA_PROJECT_B64 = "0M8R4KGxGuEAAAAAAAAAAAAAAAAAAAAAPgADAP7/CQAGAAAAAAAAAAAAAAABAAAAGgAAAAAAAAAAEAAAAQAAAAIAAAD+////AAAAAAAAAAD////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////9////AgAAAP7///8EAAAABQAAAAYAAAAHAAAACAAAAAkAAAAKAAAACwAAAAwAAAANAAAADgAAAA8AAAAQAAAAEQAAABIAAAATAAAAFAAAABUAAAAWAAAAFwAAABgAAAAZAAAA/v///xsAAAAcAAAAHQAAAP7//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////wEAAAACAAAAAwAAAAQAAAAFAAAABgAAAAcAAAAIAAAACQAAAP7///8LAAAADAAAAP7///8OAAAADwAAABAAAAARAAAAEgAAABMAAAAUAAAAFQAAABYAAAAXAAAAGAAAABkAAAAaAAAAGwAAABwAAAAdAAAA/v///x8AAAAgAAAAIQAAACIAAAAjAAAAJAAAACUAAAAmAAAAJwAAACgAAAApAAAAKgAAACsAAAAsAAAALQAAAP7///8vAAAAMAAAADEAAAAyAAAAMwAAADQAAAA1AAAANgAAADcAAAA4AAAAOQAAADoAAAA7AAAAPAAAAD0AAAD+////PwAAAEAAAABBAAAAQgAAAEMAAABEAAAARQAAAEYAAABHAAAASAAAAEkAAABKAAAASwAAAEwAAABNAAAATgAAAE8AAABQAAAAUQAAAFIAAAD+////VAAAAFUAAABWAAAAVwAAAFgAAABZAAAAWgAAAFsAAABcAAAAXQAAAF4AAABfAAAAYAAAAGEAAABiAAAA/v///2QAAABlAAAAZgAAAGcAAABoAAAAaQAAAGoAAABrAAAAbAAAAG0AAABuAAAAbwAAAHAAAABxAAAAcgAAAHMAAAB0AAAAdQAAAHYAAAB3AAAAeAAAAHkAAAB6AAAAewAAAHwAAAB9AAAAfgAAAH8AAACAAAAAgQAAAIIAAACDAAAAhAAAAIUAAACGAAAAhwAAAIgAAACJAAAAigAAAIsAAACMAAAAjQAAAI4AAACPAAAAkAAAAP7///+SAAAAkwAAAJQAAACVAAAAlgAAAJcAAACYAAAAmQAAAJoAAACbAAAAnAAAAJ0AAACeAAAAnwAAAKAAAAChAAAAogAAAKMAAACkAAAApQAAAP7///+nAAAA/v///6kAAACqAAAAqwAAAP7///+tAAAA/v///68AAACwAAAAsQAAALIAAACzAAAAtAAAALUAAAC2AAAAtwAAAP7///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////9JRD0iezhEODA3MTIyLTA2NTctNDJDOC1CQzZGLTRCNUZEMDgwMzFDOX0iDQpEb2N1bWVudD1UaGlzV29ya2Jvb2svJkgwMDAwMDAwMA0KRG9jdW1lbnQ9U2hlZXQxLyZIMDAwMDAwMDANCk1vZHVsZT1Nb2R1bGUxDQpEb2N1bWVudD1UaGlzV29ya2Jvb2sxLyZIMDAwMDAwMDANCkRvY3VtZW50PVNoZWV0Mi8mSDAwMDAwMDAwDQpOYW1lPSJWQkFQcm9qZWN0Ig0KSGVscENvbnRleHRJRD0iMCINClZlcnNpb25Db21wYXRpYmxlMzI9IjM5MzIyMjAwMCINCkNNRz0iMDcwNUJBQzJCQUMyQkVDNkJFQzZCRkM3QkZDNyINCkRQQj0iMEUwQ0IzQzlCM0M5QjZDRUI2Q0VGMTU5N0ZEQUMwIg0KR0M9IjE1MTdBOEQ0QThENUE5RDVBOUQ1Ig0KDQpbSG9zdCBFeHRlbmRlciBJbmZvXQ0KJkgwMDAwMDAwMT17MzgzMkQ2NDAtQ0Y5MC0xMUNGLThFNDMtMDBBMEM5MTEwMDVBfTtWQkU7JkgwMDAwMDAwMA0KDQpbV29ya3NwYWNlXQ0KVGhpc1dvcmtib29rPTAsIDAsIDAsIDAsIEMNClNoZWV0MT0wLCAwLCAwLCAwLCBDDQpNb2R1bGUxPTExMCwgMTQ1LCA5OTQsIDcyMSwgDQpUaGlzV29ya2Jvb2sxPTAsIDAsIDAsIDAsIEMNClNoZWV0Mj0wLCAwLCAwLCAwLCBDDQoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVGhpc1dvcmtib29rAFQAaABpAHMAVwBvAHIAawBiAG8AbwBrAAAAU2hlZXQxAFMAaABlAGUAdAAxAAAATW9kdWxlMQBNAG8AZAB1AGwAZQAxAAAAVGhpc1dvcmtib29rMQBUAGgAaQBzAFcAbwByAGsAYgBvAG8AawAxAAAAU2hlZXQyAFMAaABlAGUAdAAyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAARYBAATwAAAAxAIAANQAAACwAQAA/////wUDAAC1AwAAAAAAAAEAAADandDbAAD//wMAAAAAAAAAtgD//wEBAAAAAP////8AAAAA//8EAP//AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAAAAAMAAAAFAAAABwAAAP//////////AQEIAAAA/////3gAAAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD//wAAAABNRQAA////////AAAAAP//AAAAAP//AQEAAAAA3wD//wAAAAAIAP//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////KAAAAAAANgr/////AAAAABoI/////wAAAAACPAwA//8AAAAAAjz/////AAD//wEBAAAAAAAAAQAAAP////8BAYgAAAA4AAAA/////wKDHgL/////CAD//zAAAAAAAP///////wAAAAD//////////wAAAAAdAAAAJQAAAAsSIAL/////AAAAYAAAAAD//////////wAAAAAAAAAA//////z+SAD/////////////AAADAAMAAACEAAADAAD/////wAEAAP////+4AQAAAQABAAAAAAAAAAAAAAAAADgAAAD//////////////////////////zgAAAD/////////////////////////////////////CAAAAAAAAAAAAAAACAAEAP////8AAAAA////////////////////////////////AgAAAAAAUR9nUiAAASQAKgBcAFIAZgBmAGYAZgAqADAAYgA1ADIANgA3ADEAZgA1ADEAARAAKgBcAFIAMAAqACMAMQA0AN8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD+ygEAAwAigQgABgACAAAAAAAAgQgEHgAAABgAAAAEgQgAAgADAAgAAAD/////AQFAAAAAlgQ4AAAAAABvAP//cAAAAP////9IAAAAtgASAEhlbGxvIGZyb20gUHl0aG9uIR0AQUAiAgEAYwD/////MAAAAP////8AAAFhsABBdHRyaWJ1dABlIFZCX05hbQBlID0gIk1vZAB1bGUxIg0KUwB1YiBzYXlfaABlbGxvKCkNCgIgAABNc2dCb3gQICgiSAFUIGZyAG9tIFB5dGhvCG4hIgBERW5kIAEAbg0KAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABFgEAAfAAAADMAgAA1AAAAAACAAD/////0wIAACcDAAAAAAAAAQAAANqdhToAAP//IwEAAIgAAAC2AP//AQEAAAAA/////wAAAAD///////8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAAAAAwAAAAUAAAAHAAAA//////////8BAQgAAAD/////eAAAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP//AAAAAE1FAAD///////8AAAAA//8AAAAA//8BAQAAAADfAP//AAAAAAwA//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////8oAAAAAgBTTP////8AAAEAUxD/////AAABAFOU/////wAAAAACPP////8AAP//AQEAAAAAAQBOADAAewAwADAAMAAyADAAOAAxADkALQAwADAAMAAwAC0AMAAwADAAMAAtAEMAMAAwADAALQAwADAAMAAwADAAMAAwADAAMAAwADQANgB9AAYAAAD/////AQFAAAAAAoD+//////8QAP//KAAAAAIB//8AAAAAAAAAAP//////////AAAAAB0AAAAlAAAA/////0gAAAD/////QAAAAAAAAAAAAAEAAAAAAAAAAAD///////////////8AAAAA//////////////////////////8AAAAA//////////////////////////8AAAAAAQAAAP//////////AAAAAP///////////////////////////////wEAMAAAAFEfZ1IgAN8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD+ygEAAAD/////AQEIAAAA/////3gAAAD/////AAABqbAAQXR0cmlidXQAZSBWQl9OYW0AZSA9ICJTaGVAZXQxIg0KCuhCBGFzAnQwezAwMEAyMDgxOS0AIDAdAwhDABQCEgEkMDA0BDZ9DXxHbG9iYUJsAcRTcGFjAZJGEGFsc2UMZENyZSBhdGFibBUfUHJAZWRlY2xhAAZJImQAq1RydQ1CRXgQcG9zZRQcVGVtAHBsYXRlRGVyDGl2AiSSQnVzdG8YbWl6BEQDMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAARYBAADwAAAAxAIAANQAAAAAAgAA/////8sCAAAfAwAAAAAAAAEAAADanTJ/AAD//yMBAACIAAAAtgD//wEBAAAAAP////8AAAAA////////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAAAAAMAAAAFAAAABwAAAP//////////AQEIAAAA/////3gAAAAIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD//wAAAABNRQAA////////AAAAAP//AAAAAP//AQEAAAAA3wD//wAAAAAMAP//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////KAAAAAIAU0z/////AAABAFMQ/////wAAAQBTlP////8AAAAAAjz/////AAD//wEBAAAAAAEATgAwAHsAMAAwADAAMgAwADgAMgAwAC0AMAAwADAAMAAtADAAMAAwADAALQBDADAAMAAwAC0AMAAwADAAMAAwADAAMAAwADAAMAA0ADYAfQAAAAAA/////wEBOAAAAAKA/v//////EAD//ygAAAACAf//AAAAAAAAAAD//////////wAAAAAdAAAAJQAAAP////9IAAAAAAD//wAAAQAAAAAAAAAAAP///////////////wAAAAD//////////////////////////wAAAAD//////////////////////////wAAAAABAAAA//////////8AAAAA////////////////////////////////////////AAAAAAAA3wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP7KAQAAAP////8BAQgAAAD/////eAAAAP////8AAAGosABBdHRyaWJ1dABlIFZCX05hbQBlID0gIlNoZUBldDIiDQoK6EIEYXMCdDB7MDAwwDIwODIwLQAgBAgOQwAUAhwBJDAwNDYCfQ18R2xvYmFsIQHEU3BhYwGSRmEIbHNlDGRDcmVhEHRhYmwVH1ByZSBkZWNsYQAGSWQRAKtUcnUNQkV4cAhvc2UUHFRlbXAAbGF0ZURlcmkGdgIkkkJ1c3RvbQxpegREAzIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEWAQAB8AAAAMwCAADUAAAAAAIAAP/////TAgAAJwMAAAAAAAABAAAA2p0mIAAA//8jAQAAiAAAALYA//8BAQAAAAD/////AAAAAP///////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAAAAADAAAABQAAAAcAAAD//////////wEBCAAAAP////94AAAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA//8AAAAATUUAAP///////wAAAAD//wAAAAD//wEBAAAAAN8A//8AAAAADAD//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////ygAAAACAFNM/////wAAAQBTEP////8AAAEAU5T/////AAAAAAI8/////wAA//8BAQAAAAABAE4AMAB7ADAAMAAwADIAMAA4ADEAOQAtADAAMAAwADAALQAwADAAMAAwAC0AQwAwADAAMAAtADAAMAAwADAAMAAwADAAMAAwADAANAA2AH0ABgAAAP////8BAUAAAAACgP7//////xAA//8oAAAAAgH//wAAAAAAAAAA//////////8AAAAAHQAAACUAAAD/////SAAAAP////9AAAAAAAAAAAAAAQAAAAAAAAAAAP///////////////wAAAAD//////////////////////////wAAAAD//////////////////////////wAAAAABAAAA//////////8AAAAA////////////////////////////////AQAwAAAAUR9nUiAA3wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP7KAQAAAP////8BAQgAAAD/////eAAAAP////8AAAEFsgBBdHRyaWJ1dABlIFZCX05hbQBlID0gIlRoaQBzV29ya2JvbxBrIg0KCoxCYXMBAowwezAwMDIwUDgxOS0AEDADCEMjBRIDADQ2fQ18R2wQb2JhbAHQU3BhgmMBkkZhbHNlDCUAQ3JlYXRhYmwBFR9QcmVkZWNsEmEABklkACNUcnWBDSJFeHBvc2UUHABUZW1wbGF0ZWBEZXJpdgISkkJ1wHN0b21pegREgyMGUIAYgBwgU3ViIAGFkV9TaGVldEMAaGFuZ2UoQnkAVmFsIFNoIEEAcyBPYmplY3QELCCDCVRhcmdlCnSBC1KBFCkNCiABAABPbiBFcnJvAHIgR29UbyBTgGFmZUV4aXSDDRRJZgAkLgLLPD4gACJEYXNoYm9hgHJkIiBUaGWAGwcAFAFEBhZJbnRlcmpzADkogzUsAR8CNigAIkk3IikpIEkAcyBOb3RoaW6gZyBBbmQaGEYNDAHQHEFwcGxpY2GAdGlvbi5FbgFKwEV2ZW50c0dURCafwBBrJwEbQw/FNnNEwF2KKIsMLsBVdWUpTAtfAQDLMcMJQCXBEFNAfWFAbChZZWFykxMsQCBNb250aBUHMQNEbpQaTnVtYmVyQEZvcm1hdEHKbXBtbS15AACAwkUMRfWBXmbDA0WGScsShxwBEehBbGzCLnNEEUUQxYseOt5jg6/BDUJvAAAAAAAAAAAAAAEWAQAA8AAAAMQCAADUAAAAAAIAAP/////LAgAAHwMAAAAAAAABAAAA2p2+UAAA//8jAQAAiAAAALYA//8BAQAAAAD/////AAAAAP///////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAAAAADAAAABQAAAAcAAAD//////////wEBCAAAAP////94AAAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA//8AAAAATUUAAP///////wAAAAD//wAAAAD//wEBAAAAAN8A//8AAAAADAD//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////ygAAAACAFNM/////wAAAQBTEP////8AAAEAU5T/////AAAAAAI8/////wAA//8BAQAAAAABAE4AMAB7ADAAMAAwADIAMAA4ADEAOQAtADAAMAAwADAALQAwADAAMAAwAC0AQwAwADAAMAAtADAAMAAwADAAMAAwADAAMAAwADAANAA2AH0AAAAAAP////8BATgAAAACgP7//////xAA//8oAAAAAgH//wAAAAAAAAAA//////////8AAAAAHQAAACUAAAD/////SAAAAAAA//8AAAEAAAAAAAAAAAD///////////////8AAAAA//////////////////////////8AAAAA//////////////////////////8AAAAAAAAAAP//////////AAAAAP///////////////////////////////////////wAAAAAAAN8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD+ygEAAAD/////AQEIAAAA/////3gAAAD/////AAABsbAAQXR0cmlidXQAZSBWQl9OYW0AZSA9ICJUaGkAc1dvcmtib28gazEiDQoKkEJhAnMCkDB7MDAwMqAwODE5LQAQMAMIDkMAFAISASQwMDQ2An0NfEdsb2JhbCEB0lNwYWMBkkZhCGxzZQxkQ3JlYRB0YWJsFR9QcmUgZGVjbGEABklkEQCyVHJ1DUJFeHAIb3NlFBxUZW1wAGxhdGVEZXJpBnYCEpJCdXN0b20MaXoERAMyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADMYYUAAAEA/wkIAAAJBAAA5AQBAAAAAAAAAAAAAQAEAAIAHAEqAFwARwB7ADAAMAAwADIAMAA0AEUARgAtADAAMAAwADAALQAwADAAMAAwAC0AQwAwADAAMAAtADAAMAAwADAAMAAwADAAMAAwADAANAA2AH0AIwA0AC4AMAAjADkAIwBDADoAXABQAHIAbwBnAHIAYQBtACAARgBpAGwAZQBzAFwAQwBvAG0AbQBvAG4AIABGAGkAbABlAHMAXABNAGkAYwByAG8AcwBvAGYAdAAgAFMAaABhAHIAZQBkAFwAVgBCAEEAXABWAEIAQQA2AFwAVgBCAEUANgAuAEQATABMACMAVgBpAHMAdQBhAGwAIABCAGEAcwBpAGMAIABGAG8AcgAgAEEAcABwAGwAaQBjAGEAdABpAG8AbgBzAAAAAAAAAAAAAAAAABoBKgBcAEcAewAwADAAMAAyADAAOAAxADMALQAwADAAMAAwAC0AMAAwADAAMAAtAEMAMAAwADAALQAwADAAMAAwADAAMAAwADAAMAAwADQANgB9ACMAMQAuADYAIwAwACMAQwA6AFwAcAByAG8AZwByAGEAbQAgAGYAaQBsAGUAcwBcAG0AaQBjAHIAbwBzAG8AZgB0ACAAbwBmAGYAaQBjAGUAIAAyADAAMAA3AFwATwBmAGYAaQBjAGUAMQAyAFwARQBYAEMARQBMAC4ARQBYAEUAIwBNAGkAYwByAG8AcwBvAGYAdAAgAEUAeABjAGUAbAAgADEAMgAuADAAIABPAGIAagBlAGMAdAAgAEwAaQBiAHIAYQByAHkAAAAAAAAAAAAAAAAAvAAqAFwARwB7ADAAMAAwADIAMAA0ADMAMAAtADAAMAAwADAALQAwADAAMAAwAC0AQwAwADAAMAAtADAAMAAwADAAMAAwADAAMAAwADAANAA2AH0AIwAyAC4AMAAjADAAIwBDADoAXABXAEkATgBEAE8AVwBTAFwAcwB5AHMAdABlAG0AMwAyAFwAcwB0AGQAbwBsAGUAMgAuAHQAbABiACMATwBMAEUAIABBAHUAdABvAG0AYQB0AGkAbwBuAAAAAAAAAAAAAAAAACgBKgBcAEcAewAyAEQARgA4AEQAMAA0AEMALQA1AEIARgBBAC0AMQAwADEAQgAtAEIARABFADUALQAwADAAQQBBADAAMAA0ADQARABFADUAMgB9ACMAMgAuADQAIwAwACMAQwA6AFwAUAByAG8AZwByAGEAbQAgAEYAaQBsAGUAcwBcAEMAbwBtAG0AbwBuACAARgBpAGwAZQBzAFwATQBpAGMAcgBvAHMAbwBmAHQAIABTAGgAYQByAGUAZABcAE8ARgBGAEkAQwBFADEAMgBcAE0AUwBPAC4ARABMAEwAIwBNAGkAYwByAG8AcwBvAGYAdAAgAE8AZgBmAGkAYwBlACAAMQAyAC4AMAAgAE8AYgBqAGUAYwB0ACAATABpAGIAcgBhAHIAeQAAAAAAAAAAAAAAAAAFAAIAAgABAAIAAgAEAAQCAAAGAgEACAIAAAoCAQAQAv///////wAAAAD//wAAUR9nUiAAAQD//wIAAAD//////////wQAAwD//////////////////////////////////////////////////////////wEAAAAAAAAAAAAAAAAAAAAAAAAA2p0FABgAVABoAGkAcwBXAG8AcgBrAGIAbwBvAGsAFAAwAF0ANQAyADYANwAxAGYAMgAxAP//FQIYAFQAaABpAHMAVwBvAHIAawBiAG8AbwBrAP//JiAAAAAAAAAAAgAAAC0DAAD//wwAUwBoAGUAZQB0ADEAFAAwAF4ANQAyADYANwAxAGYAMgAxAP//GQIMAFMAaABlAGUAdAAxAP//hToAAAAAAAAYAgAAAC0DAAD//w4ATQBvAGQAdQBsAGUAMQAUADAAYgA1ADIANgA3ADEAZgA1ADEA//8eAg4ATQBvAGQAdQBsAGUAMQD//9DbAAAAAAAAMAIAAAC7AwAA//8aAFQAaABpAHMAVwBvAHIAawBiAG8AbwBrADEAFAAwADUANQA2ADEAMQBjADAAYgBlAP//JQIaAFQAaABpAHMAVwBvAHIAawBiAG8AbwBrADEA//++UAAAAAAAAFACAAAAJQMAAP//DABTAGgAZQBlAHQAMgAUADAANgA1ADYAMQAxAGMAMABiAGUA//8nAgwAUwBoAGUAZQB0ADIA//8yfwAAAAAAAGgCAAAAJQMAAP///////wEBiAIAABgCAAD/////////////////////aAIAAP////////////////////////////////////////////////////////////////////////////////////////////////////8wAgAA////////////////////////////////AAIAAP///////////////////////////////////////////////////////////////////////////////////////////////1ACAAD/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////rNztGtDOSEyMTMXxwDvwFP////8BAAAAGt2hIqKi/km0kB0c45MtUv////8BAAAAU4biDKx9xkiedsKr70dxO/////8BAAAA/////zAAAAD/A1s5f1x+RIZ2wPAlw/Xp/////wEAAAB/NCo06yLUQaTZXrMNyxhP/////wEAAAD/////GAIAAIAAAAAAABMBFAD/ALgmAAAFBEV4Y2VsgCsQAAMEVkJB9+IQAAUEV2luMTbBfhAABQRXaW4zMgd/EAADBE1hY7OyEAAEBFZCQTatIxAACARQcm9qZWN0MQoXEAAGBHN0ZG9sZZNgEAAKBFZCQVByb2plY3S+vxAABgRPZmZpY2UVdRAADARUaGlzV29ya2Jvb2t84xAACYAAAP8DAQBfRXZhbHVhdGUY2RAABgRTaGVldDHoGhAACARXb3JrYm9va2sYEAAJBFdvcmtzaGVldMH+EAAHBE1vZHVsZTFiERAACYQIAP8D//9zYXlfaGVsbG8wdBAABgBNc2dCb3iXUhAADQRUaGlzV29ya2Jvb2sxNV0QAAYEU2hlZXQy6RoQAAL//wEBeAAAAP///////////////////////w4CAgD//xAC/////xICAwD//xUCAAAOAP///////xkCAQAGAP///////////////x8CAgALAP///////////////yUCAwALACcCBAAIAAACAQD//wICAAD//////////////////w8AFAAAAAEAWgAAAAAAAAAAAAAAAAAAAAAAAAABAAAAAAAAAAAAAAAAAAAAAAABAAAAAAAAAAEAAAAAAAAAAAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAk0sqhQEAEAAAAP//AAAAAAEAAgD//wAAAAABAAAAAgAAAAAAAQACAAIAAAAAAAEABQAFAAUABQAFAAUABQAFAAUABQAFAAUAAAByVQABAACAAAAAgAAAAIAAAAAEAAB+AQAAfgEAAH4BAAB+AQAAfgIAAH4DAAB+awAAfwAAAAAVAAAACQAAAAAAAQAIAAAAAAAAABEBAAAAAAAAJwK+MGGQ0Uu6Rawc6DExiQEACQQAAAkIAADkBAAAAAAAAAEA/////wUAAgoAAP///////////////wAAAAC5AAAAAAAAAAIKAAD///////////////8AAAAA4QAAAAAAAAABCEEACQAAAAAAAgChBQAAAAAAAP////9wAAAAAAAAAP////8BAIEFAAAAAAAA//8AAPkAAAAAAAAAAwoEAP///////////////wAAAAD5BgAAAAAAAAMKBAD///////////////8AAAAAIQcAAAAAAAAEANEBAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOkCAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAkEAAAAAAAAAAAAAAAAAAAAAAAAAAAAACkFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAAAIMAAAAVGhpc1dvcmtib29rAgAAAgYAAABTaGVldDECAAACBwAAAE1vZHVsZTEDAAACCgAAAFZCQVByb2plY3QEAAAD7wQCAAAAAADAAAAAAAAARgwAAAIvAAAAQzpcUFJPR1JBfjFcQ09NTU9OfjFcTUlDUk9TfjFcVkJBXFZCQTZcVkJFNi5ETEwBAAACAwAAAFZCQQoAAAoxAQAAAAAAAP////8EAAAACQAAAFkBAAAAAAAAwQEAAAAAAAAwAAAAAAAAAAAAAAAAAAQAAAMTCAIAAAAAAMAAAAAAAABGDwAAAjkAAABDOlxwcm9ncmFtIGZpbGVzXG1pY3Jvc29mdCBvZmZpY2UgMjAwN1xPZmZpY2UxMlxFWENFTC5FWEUCAAACBQAAAEV4Y2VsCgAACikCAAAAAAAA/////wEABgAAAAAAUQIAAAAAAADRAgAAAAAAAEAAAAAAAAAAAAAAAAAABAAAAzAEAgAAAAAAwAAAAAAAAEYIAAACHwAAAEM6XFdJTkRPV1Ncc3lzdGVtMzJcc3Rkb2xlMi50bGICAAACBgAAAHN0ZG9sZQcAAH8KAAAKQQMAAAAAAAD/////AgAAAAAAAABpAwAAAAAAALEDAAAAAAAAUAAAAAAAAAAAAAAAAAAEAAADTND4LfpbGxC95QCqAETeUhAAAAI/AAAAQzpcUHJvZ3JhbSBGaWxlc1xDb21tb24gRmlsZXNcTWljcm9zb2Z0IFNoYXJlZFxPRkZJQ0UxMlxNU08uRExMAgAAAgYAAABPZmZpY2UKAAAKYQQAAAAAAAD/////AgAEAAAAAACJBAAAAAAAABEFAAAAAAAAYAAAAAAAAAAAAAAAAAADAAACCQAAAHNheV9oZWxsbwMAAA0MAAwACAAAAAAAAAAAAAkAAAseAAAASABlAGwAbABvACAAZgByAG8AbQAgAFAAZQByAGwAAwAAAggAAABWQkU2LkRMTA0AAAcRBgAAAAAAAP////9TAgsAiAAAAAAAAAAKAAALIAAAAEgAZQBsAGwAbwAgAGYAcgBvAG0AIABQAGUAcgBsACEABAAAAg0AAABUaGlzV29ya2Jvb2sxAgAAAgYAAABTaGVldDIZAAB/AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAclWAAAAAAAAAAIAAAACAAAAAAAAAAAoAAAAJAAAAAAAAAP///////////////wAAAAD//////////wkAAAAAAAMA//////////90AAB/AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAByVYAAAACAAAAAgAAAAIAAAAACAAB+fAAAfwAAAAAOAAAACQAAAAAAAAAJAAAAAAADAAgAAAAAAAIAAQABAAMAAADBBQAAAAAAADEGAAAAAAAAoQYAAAAAAABwAAAAMAAAAAACACsn/P4nHP8nPP/1AAAAADps/wIATlz/BFz/CgEAFAA2CABc/zz/HP/8/gAAFAAAAAAEAIAAMAAkAAgAAAAAABMAAAAAAAwAAAAAAAAAAAAAABwAAAAAAAQAAAAAAFz/AgA8/wIAHP8CAPz+AgAEAAASAABrAAB/AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAclWAAAAAAAAAAIAAAACAAAAAAAAAABAAAAAJAAAAAAACAP//////////AAAAAAgAAAAEACQAgQAAAAAAAgAAAABgAAD9//////////////8AAAAAAAAAAAAAAAAAAABuAAB/AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABZrKAAQAEAAAAAQAwKgICkAkAcBQGSAMAggIAZOQEBAAKABwAVkJBUHJvamWIY3QFADQAAEACFGoGAgo9AgoHAnIBFAgFBhIJAhJRH2dSIJQADAJKPAIKFgABcoBzdGRvbGU+AhkAcwB0AGQAbwCAbABlAA0AaAAlAl4AAypcR3swMIAwMjA0MzAtAAgdBARDAAoCDgESMDA0ADZ9IzIuMCMwACNDOlxXSU5EAE9XU1xzeXN0IGVtMzJcA2UyLgB0bGIjT0xFIABBdXRvbWF0aRxvbgBgAAGDRU9mZkRpY4RFTwBmgABp1ABjgkWegBGUgAGBRQAyREY4RDA0QwAtNUJGQS0xMIAxQi1CREU1gEXUQUGAQzSABTKIRYCYAGdyYW0gRmlsAGVzXENvbW1vAm4EBk1pY3JvcwBvZnQgU2hhcgBlZFxPRkZJQwBFMTJcTVNPLjBETEwjhxCDTSAxEYBxIE9igcEgTGlgYnJhcnkASwABDxGC1AUAE4ID2p0ZAYKoVGhpc1dvcgBrYm9va0cAGAWAE1SAq2kAcwBXQYCzcgBrAGLAAW9VwAEazgsy2gscwBIAFABIQgExQngtAwAsAB5CAgEFLMIhJiAqIkIIK0IBGUJ8U2igZWV0MUfCG1NAI4JlQFh0ADEAGkgHpjJOB+MbhTrLGwfAHUBNb2R1bGUAHA7jAQOAO2QAdYKYgRwICCwyAA8ITx27UDnQ29IhgBYAAEU5DUADSV69wB4agQSUXsEhjgwywEOkAABVazEATyYlUCb0vlDUXzLNX+EN4XGiM5oy7S8y/w3iPTJ/6Q0CEIAeAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgBvAG8AdAAgAEUAbgB0AHIAeQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABYABQH//////////wEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADAAAAAC4AAAAAAABQAFIATwBKAEUAQwBUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAACAQMAAAACAAAA/////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSAgAAAAAAAFAAUgBPAEoARQBDAFQAdwBtAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAAIA////////////////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACgAAAJUAAAAAAAAAVgBCAEEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgAAQD//////////wUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABfAF8AUwBSAFAAXwAwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAACAQoAAAD//////////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJEAAAAjBQAAAAAAAF8AXwBTAFIAUABfADEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAAIBDAAAAA0AAAD/////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAApgAAAFIAAAAAAAAAXwBfAFMAUgBQAF8AMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAAAgD///////////////8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACoAAAA4AAAAAAAAABfAF8AUwBSAFAAXwAzAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAACAQYAAAD//////////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKwAAABnAAAAAAAAAF8AVgBCAEEAXwBQAFIATwBKAEUAQwBUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAaAAIA////////////////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYwAAAEgLAAAAAAAAZABpAHIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgAAgD///////////////8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACuAAAAagIAAAAAAABNAG8AZAB1AGwAZQAxAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAACAP///////////////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA0AAAAgBAAAAAAAAFMAaABlAGUAdAAxAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOAAIBCQAAAP//////////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHgAAANoDAAAAAAAAUwBoAGUAZQB0ADIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA4AAgALAAAABAAAAP////8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAuAAAA0QMAAAAAAABUAGgAaQBzAFcAbwByAGsAYgBvAG8AawAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGgACAAcAAAAOAAAA/////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD4AAAA2BQAAAAAAAFQAaABpAHMAVwBvAHIAawBiAG8AbwBrADEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAcAAIBCAAAAP//////////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUwAAANoDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA="

BRANCH_NAMES = {
    "101": "HEAD OFFICE",
    "102": "AL KHAIR",
    "104": "NOORA",
    "107": "MAGNUS YAQOOT",
    "109": "MAGNUS HAMRA",
    "111": "MAGNUS EXPRESS",
    "112": "MAGNUS HADA SHAM",
}

ALIASES = {
    "branch": ["branch"],
    "date": ["date"],
    "sales_target": ["sales target"],
    "sales_achieved": ["sales achivement", "sales achievement", "sales achieved"],
    "nob_target": ["nob target"],
    "nob_achieved": ["nob achievemnet", "nob achievement", "nob achieved"],
    "abv_target": ["abv target"],
    "abv_actual": ["abv achievement", "abv achieved"],
    "gp_target": ["gp target"],
    "gp_actual": ["gp acheivment", "gp achievement", "gp achieved"],
}


def clean_header_for_ui(value):
    text = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _extract_source_info(input_file):
    if not input_file:
        return [], []
    book = pd.ExcelFile(input_file, engine="openpyxl")
    controlled_sheet = next((s for s in ["Data Update", "Data"] if s in book.sheet_names), None)
    months = set()
    branches = set()

    def read_frame(sheet_name, fixed_branch=None):
        raw = pd.read_excel(input_file, sheet_name=sheet_name, engine="openpyxl")
        raw.columns = [clean_header_for_ui(c) for c in raw.columns]
        date_col = next((c for c in ALIASES["date"] if c in raw.columns), None)
        branch_col = next((c for c in ALIASES["branch"] if c in raw.columns), None)
        if date_col:
            dt = pd.to_datetime(raw[date_col], errors="coerce").dropna()
            months.update(dt.dt.to_period("M").tolist())
        if fixed_branch is not None:
            branches.add(str(fixed_branch).strip())
        elif branch_col:
            for value in raw[branch_col].dropna().astype(str):
                value = value.strip()
                if value and value.lower() != "nan":
                    branches.add(value)

    if controlled_sheet:
        read_frame(controlled_sheet)
        try:
            branch_master = pd.read_excel(input_file, sheet_name=controlled_sheet, usecols="X", engine="openpyxl")
            if len(branch_master.columns):
                for value in branch_master.iloc[:, 0].dropna().astype(str):
                    value = value.strip()
                    if value and value != "All Branches":
                        branches.add(value)
        except Exception:
            pass
    else:
        for sheet in [s for s in book.sheet_names if str(s).strip().isdigit()]:
            read_frame(sheet, fixed_branch=sheet)
        branches = {f"{b} - {BRANCH_NAMES.get(str(b), 'NAME NOT MAPPED')}" for b in branches}

    month_labels = []
    if months:
        latest = max(months)
        options = sorted(months | {latest + 1})
        month_labels = [p.to_timestamp().strftime("%b-%Y") for p in options]
    return month_labels, sorted(branches)


def inspect_upload(uploaded_file):
    if not uploaded_file:
        return gr.Dropdown(choices=[], value=None), gr.Dropdown(choices=[], value=None), "Upload an Excel source file first."
    try:
        months, branches = _extract_source_info(uploaded_file)
        if not months:
            raise ValueError("No valid Date column/months were found in the workbook.")
        # Default to latest source month, not the auto-added next month.
        default_month = months[-2] if len(months) >= 2 else months[-1]
        default_model = branches[0] if branches else None
        msg = f"Source loaded: {len(branches)} branch(es). Choose the controlled month/status and generate."
        return gr.Dropdown(choices=months, value=default_month), gr.Dropdown(choices=branches, value=default_model), msg
    except Exception as exc:
        return gr.Dropdown(choices=[], value=None), gr.Dropdown(choices=[], value=None), f"Could not read the source file: {exc}"


def _build_dashboard_core(input_file, selected_month, selected_status, add_branch, new_branch, model_branch):
    if not input_file:
        raise gr.Error("Please upload the source Excel file.")
    if not selected_month:
        raise gr.Error("Please select the controlled month.")

    selected_status = (selected_status or "OPEN").strip().upper()
    if selected_status not in {"OPEN", "CLOSED"}:
        raise gr.Error("Period status must be OPEN or CLOSED.")

    add_branch_value = "YES" if str(add_branch).strip().upper() == "YES" else "NO"
    if add_branch_value == "YES":
        if not str(new_branch or "").strip():
            raise gr.Error("Enter the new branch as 000 - BRANCH NAME.")
        if not str(model_branch or "").strip():
            raise gr.Error("Choose an existing model branch.")

    INPUT_FILE = str(input_file)
    OPEN_MONTH = str(selected_month).strip()
    PERIOD_STATUS = selected_status
    ADD_BRANCH = add_branch_value
    REQUESTED_NEW_BRANCH = str(new_branch or "").strip()
    REQUESTED_MODEL_BRANCH = str(model_branch or "").strip()
    RUNNING_IN_COLAB = False
    FORCE_PERIOD_PROMPT = False

    output_dir = Path(tempfile.mkdtemp(prefix="mtd_dashboard_"))
    OUTPUT_FILE = str(output_dir / "MTD_Dynamic_Dashboard_Date_Mode_Fixed.xlsm")

    logs = io.StringIO()
    try:
        with contextlib.redirect_stdout(logs):
            # @title
            def clean_header(value):
                text = "" if value is None else str(value)
                return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()

            def numeric(series):
                return pd.to_numeric(series, errors="coerce")

            def protect_workbook_structure(path):
                """Lock sheet structure without rewriting the workbook or its VBA project."""
                source = Path(path).read_bytes()
                output = io.BytesIO()
                protection = (
                    f'<workbookProtection workbookPassword="{hash_password(PASSWORD)}" '
                    'lockStructure="1"/>'
                ).encode("utf-8")
                with zipfile.ZipFile(io.BytesIO(source), "r") as zin, zipfile.ZipFile(output, "w") as zout:
                    for item in zin.infolist():
                        payload = zin.read(item.filename)
                        if item.filename == "xl/workbook.xml":
                            if b"<workbookProtection " in payload:
                                payload = re.sub(
                                    rb"<workbookProtection\b[^>]*/>",
                                    protection,
                                    payload,
                                    count=1,
                                )
                            else:
                                payload = payload.replace(b"<bookViews>", protection + b"<bookViews>", 1)
                        zout.writestr(item, payload)
                Path(path).write_bytes(output.getvalue())

            book = pd.ExcelFile(INPUT_FILE, engine="openpyxl")
            records = []
            controlled_sheet = next((s for s in ["Data Update", "Data"] if s in book.sheet_names), None)
            persisted_branches = []

            aliases = {
                "branch": ["branch"],
                "date": ["date"],
                "sales_target": ["sales target"],
                "sales_achieved": ["sales achivement", "sales achievement", "sales achieved"],
                "nob_target": ["nob target"],
                "nob_achieved": ["nob achievemnet", "nob achievement", "nob achieved"],
                "abv_target": ["abv target"],
                "abv_actual": ["abv achievement", "abv achieved"],
                "gp_target": ["gp target"],
                "gp_actual": ["gp acheivment", "gp achievement", "gp achieved"],
            }

            def standardize_source(raw, fixed_branch=None):
                raw = raw.copy()
                raw.columns = [clean_header(c) for c in raw.columns]
                found = {key: next((c for c in choices if c in raw.columns), None)
                         for key, choices in aliases.items()}
                if not found["date"] or not found["sales_achieved"]:
                    return None
                if fixed_branch is None and not found["branch"]:
                    return None
                df = pd.DataFrame({
                    "Branch": str(fixed_branch).strip() if fixed_branch is not None else raw[found["branch"]].astype(str).str.strip(),
                    "Date": pd.to_datetime(raw[found["date"]], errors="coerce"),
                })
                for key, label in [
                    ("sales_target", "Sales Target"), ("sales_achieved", "Sales Achieved"),
                    ("nob_target", "NOB Target"), ("nob_achieved", "NOB Achieved"),
                    ("abv_target", "ABV Target"), ("abv_actual", "ABV Actual"),
                    ("gp_target", "GP Target"), ("gp_actual", "GP Actual")]:
                    df[label] = numeric(raw[found[key]]) if found[key] else np.nan
                return df[df["Date"].notna() & df["Sales Achieved"].notna()].copy()

            if controlled_sheet:
                raw = pd.read_excel(INPUT_FILE, sheet_name=controlled_sheet, engine="openpyxl")
                controlled = standardize_source(raw)
                if controlled is not None:
                    records.append(controlled)
                    print(f"Loaded prior controlled file from {controlled_sheet!r}.")
                try:
                    branch_master = pd.read_excel(INPUT_FILE, sheet_name=controlled_sheet, usecols="X", engine="openpyxl")
                    if len(branch_master.columns):
                        persisted_branches = [
                            str(value).strip() for value in branch_master.iloc[:, 0].dropna().tolist()
                            if str(value).strip() and str(value).strip() != "All Branches"
                        ]
                except Exception:
                    persisted_branches = []
            else:
                branch_sheets = [s for s in book.sheet_names if str(s).strip().isdigit()]
                for sheet in branch_sheets:
                    raw = pd.read_excel(INPUT_FILE, sheet_name=sheet, engine="openpyxl")
                    branch_frame = standardize_source(raw, fixed_branch=sheet)
                    if branch_frame is None:
                        print(f"Skipped sheet {sheet!r}: required Date/Sales columns not found")
                        continue
                    records.append(branch_frame)

            if not records:
                raise ValueError("No usable daily branch data found in the workbook.")

            data = pd.concat(records, ignore_index=True).sort_values(["Date", "Branch"])
            BRANCH_NAMES = {
                "101": "HEAD OFFICE",
                "102": "AL KHAIR",
                "104": "NOORA",
                "107": "MAGNUS YAQOOT",
                "109": "MAGNUS HAMRA",
                "111": "MAGNUS EXPRESS",
                "112": "MAGNUS HADA SHAM",
            }
            if not controlled_sheet:
                data["Branch"] = data["Branch"].map(
                    lambda code: f"{code} - {BRANCH_NAMES.get(str(code), 'NAME NOT MAPPED')}"
                )
            data["Day"] = data["Date"].dt.strftime("%a")
            data = data[["Branch", "Date", "Day", "Sales Target", "Sales Achieved",
                         "NOB Target", "NOB Achieved", "ABV Target", "ABV Actual",
                         "GP Target", "GP Actual"]]

            branches = sorted(set(data["Branch"].unique().tolist() + persisted_branches))
            dates = sorted(pd.to_datetime(data["Date"].dropna().unique()).tolist())
            source_months = sorted(pd.to_datetime(data["Date"]).dt.to_period("M").unique().tolist())

            def request_period_control(source_periods):
                """Ask the owner which month to control and whether it is open or closed."""
                selected_month = OPEN_MONTH
                selected_status = PERIOD_STATUS
                prompt_enabled = RUNNING_IN_COLAB or FORCE_PERIOD_PROMPT
                latest_source = max(source_periods)
                month_options = sorted(set(source_periods + [latest_source + 1]))

                if prompt_enabled and not selected_month:
                    print("\nMONTH CONTROL")
                    print("Choose the month for the new controlled dashboard:")
                    for number, period in enumerate(month_options, start=1):
                        note = ""
                        if period == latest_source:
                            note = "  [latest source month]"
                        elif period == latest_source + 1:
                            note = "  [next month]"
                        print(f"  {number}. {period.to_timestamp():%b-%Y}{note}")
                    default_number = month_options.index(latest_source) + 1
                    while True:
                        choice = input(f"Enter month number [{default_number}]: ").strip()
                        if not choice:
                            choice = str(default_number)
                        if choice.isdigit() and 1 <= int(choice) <= len(month_options):
                            selected_month = month_options[int(choice)-1].to_timestamp().strftime("%b-%Y")
                            break
                        print("Please enter one of the month numbers shown above.")

                if prompt_enabled and not selected_status:
                    print("\nChoose the period status:")
                    print("  1. OPEN   - selected-month rows are highlighted for update")
                    print("  2. CLOSED - period is marked closed; data-entry cells remain editable")
                    while True:
                        choice = input("Enter status number [1]: ").strip() or "1"
                        choice = choice.upper()
                        if choice in {"1", "OPEN"}:
                            selected_status = "OPEN"
                            break
                        if choice in {"2", "CLOSED"}:
                            selected_status = "CLOSED"
                            break
                        print("Please enter 1/OPEN or 2/CLOSED.")

                selected_month = selected_month or latest_source.to_timestamp().strftime("%b-%Y")
                selected_status = selected_status or "OPEN"
                if selected_status not in {"OPEN", "CLOSED"}:
                    raise ValueError("PERIOD_STATUS must be OPEN or CLOSED.")
                print(f"\nSelected control: {selected_month} | {selected_status}")
                return selected_month, selected_status

            OPEN_MONTH, PERIOD_STATUS = request_period_control(source_months)
            if OPEN_MONTH:
                try:
                    latest_period = pd.Period(pd.to_datetime(OPEN_MONTH, format="%b-%Y"), freq="M")
                except ValueError as exc:
                    raise ValueError("OPEN_MONTH must use the format Mon-YYYY, for example Sep-2026.") from exc
            else:
                latest_period = max(source_months)

            def normalize_branch_name(value):
                """Return the standard '000 - BRANCH NAME' format used by the dashboard."""
                text = re.sub(r"\s+", " ", str(value).strip())
                match = re.fullmatch(r"(\d{3})\s*-\s*(.+)", text)
                if not match or not match.group(2).strip():
                    return None
                return f"{match.group(1)} - {match.group(2).strip().upper()}"

            def request_branch_control(existing_branches, model_branches):
                """Ask whether to add one approved branch and which existing branch supplies its targets."""
                prompt_enabled = RUNNING_IN_COLAB or FORCE_PERIOD_PROMPT
                requested = ADD_BRANCH
                new_branch = REQUESTED_NEW_BRANCH
                model_branch = REQUESTED_MODEL_BRANCH

                if prompt_enabled and not requested:
                    print("\nBRANCH CONTROL")
                    print("Add a new branch to the approved branch list?")
                    print("  1. YES - add a new branch using an existing branch model")
                    print("  2. NO  - keep the current branch list")
                    while True:
                        choice = (input("Enter option [2]: ").strip() or "2").upper()
                        if choice in {"1", "YES", "Y"}:
                            requested = "YES"
                            break
                        if choice in {"2", "NO", "N"}:
                            requested = "NO"
                            break
                        print("Please enter 1/YES or 2/NO.")

                requested = requested or "NO"
                if requested in {"1", "YES", "Y"}:
                    requested = "YES"
                elif requested in {"2", "NO", "N"}:
                    requested = "NO"
                else:
                    raise ValueError("MTD_ADD_BRANCH must be YES or NO.")

                if requested == "NO":
                    print("Branch control: no new branch requested.")
                    return None, None

                if not model_branches:
                    raise ValueError("A new branch cannot be added because no existing branch model is available.")

                while True:
                    if not new_branch:
                        new_branch = input("Enter new branch code and name (example 115 - NEW BRANCH): ").strip()
                    normalized = normalize_branch_name(new_branch)
                    if normalized and normalized.casefold() not in {b.casefold() for b in existing_branches}:
                        new_branch = normalized
                        break
                    message = "Use the format 000 - BRANCH NAME and enter a branch that does not already exist."
                    if not prompt_enabled:
                        raise ValueError(message)
                    print(message)
                    new_branch = ""

                if prompt_enabled and not model_branch:
                    print("\nChoose the existing branch model for Sales Target, NOB Target, ABV Target and GP Target:")
                    for number, branch in enumerate(model_branches, start=1):
                        print(f"  {number}. {branch}")
                    while True:
                        choice = input("Enter model branch number [1]: ").strip() or "1"
                        if choice.isdigit() and 1 <= int(choice) <= len(model_branches):
                            model_branch = model_branches[int(choice)-1]
                            break
                        print("Please enter one of the branch numbers shown above.")
                elif model_branch.isdigit() and 1 <= int(model_branch) <= len(model_branches):
                    model_branch = model_branches[int(model_branch)-1]

                matched_model = next((b for b in model_branches if b.casefold() == model_branch.casefold()), None)
                if not matched_model:
                    raise ValueError("MTD_MODEL_BRANCH must match an existing branch name or its displayed number.")

                print(f"New approved branch: {new_branch}")
                print(f"Existing branch model: {matched_model}")
                return new_branch, matched_model

            model_branches = sorted(data["Branch"].dropna().unique().tolist())
            NEW_BRANCH, MODEL_BRANCH = request_branch_control(branches, model_branches)
            if NEW_BRANCH:
                branches = sorted(set(branches + [NEW_BRANCH]))

            new_branch_template_rows = []
            if NEW_BRANCH and PERIOD_STATUS == "OPEN":
                model_rows = data[data["Branch"] == MODEL_BRANCH].sort_values("Date")
                target_columns = ["Sales Target", "NOB Target", "ABV Target", "GP Target"]
                target_values = {}
                for column in target_columns:
                    available = model_rows[column].dropna()
                    target_values[column] = float(available.iloc[-1]) if len(available) else np.nan
                for _ in range(int(latest_period.days_in_month)):
                    new_branch_template_rows.append({
                        "Branch": NEW_BRANCH,
                        "Date": pd.NaT,
                        "Day": "",
                        "Sales Target": target_values["Sales Target"],
                        "Sales Achieved": np.nan,
                        "NOB Target": target_values["NOB Target"],
                        "NOB Achieved": np.nan,
                        "ABV Target": target_values["ABV Target"],
                        "ABV Actual": np.nan,
                        "GP Target": target_values["GP Target"],
                        "GP Actual": np.nan,
                        "_new_branch_placeholder": True,
                    })
                print(f"Prepared {len(new_branch_template_rows)} blank daily rows for {NEW_BRANCH}; target settings copied from {MODEL_BRANCH}.")

            months = sorted(set(source_months + [latest_period]))
            open_month_start = latest_period.to_timestamp()
            open_month_end = latest_period.to_timestamp(how="end").normalize()
            date_options = sorted(set(dates + pd.date_range(open_month_start, open_month_end).tolist()))
            default_branch = "All Branches"
            default_month = latest_period.to_timestamp().strftime("%b-%Y")
            default_date_filter = "All Dates"
            default_date = pd.Timestamp(max(dates))
            print(f"Loaded {len(data):,} daily rows | {len(branches)} branches | through {default_date:%d-%b-%Y}")
            print(f"Controlled period: {default_month} | Status: {PERIOD_STATUS}")



            # @title
            def subset(branch, month="All Months", date_filter="All Dates"):
                frame = data.copy()
                if branch != "All Branches":
                    frame = frame[frame["Branch"] == branch]
                if month != "All Months":
                    frame = frame[frame["Date"].dt.strftime("%b-%Y") == month]
                if date_filter != "All Dates":
                    frame = frame[frame["Date"] == pd.Timestamp(date_filter)]
                return frame

            current = subset(default_branch, default_month, default_date_filter)
            previous_period = latest_period - 1
            previous = subset(default_branch, previous_period.to_timestamp().strftime("%b-%Y"), "All Dates")
            current_last_day = int(current["Date"].max().day) if len(current) else 0
            previous_month_end = previous_period.to_timestamp(how="end").normalize()
            previous_cutoff = previous_period.to_timestamp() + pd.Timedelta(days=max(current_last_day - 1, 0))
            previous = previous[previous["Date"] <= min(previous_cutoff, previous_month_end)]
            def safe_total(series):
                value = series.sum(min_count=1)
                return 0.0 if pd.isna(value) else float(value)

            sales_target = safe_total(current["Sales Target"])
            sales_actual = safe_total(current["Sales Achieved"])
            previous_sales = safe_total(previous["Sales Achieved"])
            period_change = (sales_actual - previous_sales) / previous_sales if previous_sales else 0
            nob_target = safe_total(current["NOB Target"])
            nob_actual = safe_total(current["NOB Achieved"])
            sales_pct = sales_actual / sales_target if sales_target else 0
            nob_pct = nob_actual / nob_target if nob_target else 0
            abv_actual = sales_actual / nob_actual if nob_actual else 0
            abv_target = current["ABV Target"].dropna().mean() if current["ABV Target"].notna().any() else 0
            variance = sales_actual - sales_target
            elapsed_days = current_last_day
            days_in_month = int(latest_period.days_in_month)
            days_remaining = max(days_in_month - elapsed_days, 0)
            month_end_forecast = sales_actual / elapsed_days * days_in_month if elapsed_days else 0
            projected_month_target = sales_target / elapsed_days * days_in_month if elapsed_days else 0
            forecast_gap = month_end_forecast - projected_month_target
            remaining_target = max(projected_month_target - sales_actual, 0)
            required_daily_sales = remaining_target / days_remaining if days_remaining else 0

            def insight_lines(frame):
                lines = []
                by_branch = frame.groupby("Branch", as_index=False).agg(
                    Target=("Sales Target", "sum"), Actual=("Sales Achieved", "sum"),
                    NOB_Target=("NOB Target", "sum"), NOB_Actual=("NOB Achieved", "sum"))
                by_branch["Pct"] = np.where(by_branch["Target"] != 0, by_branch["Actual"] / by_branch["Target"], 0)
                by_branch["NOB_Pct"] = np.where(by_branch["NOB_Target"] != 0, by_branch["NOB_Actual"] / by_branch["NOB_Target"], 0)
                for _, item in by_branch.sort_values("Pct").iterrows():
                    pct, bills = float(item["Pct"]), float(item["NOB_Pct"])
                    symbol = "▲ POSITIVE" if pct >= 1 else ("● WATCH" if pct >= 0.85 else "⚠ CRITICAL")
                    gap = float(item["Actual"] - item["Target"])
                    gap_word = "target exceeded" if gap >= 0 else "target gap"
                    action = "Maintain the current pace." if pct >= 1 else ("Close the remaining gap and review ABV." if pct >= 0.85 else "Start a daily recovery plan and review conversion.")
                    lines.append((symbol, f"{item['Branch']} — sales {pct:.1%}; {gap_word} SAR {abs(gap):,.0f}; bills {bills:.1%}. {action}"))
                if not len(lines):
                    lines.append(("ℹ DATA", "No valid records are available for the selected period."))
                return lines[:6]

            insights = insight_lines(current)
            default_branch_summary = current.groupby("Branch", as_index=False).agg(Target=("Sales Target","sum"), Actual=("Sales Achieved","sum"))
            default_branch_summary["Pct"] = np.where(default_branch_summary["Target"] != 0, default_branch_summary["Actual"] / default_branch_summary["Target"], 0)
            best_default = ""
            if len(default_branch_summary):
                best_row = default_branch_summary.sort_values("Pct", ascending=False).iloc[0]
                best_default = f"{best_row['Branch']} at {best_row['Pct']:.1%}"



            # @title
            excel_output = io.BytesIO()
            with pd.ExcelWriter(excel_output, engine="xlsxwriter", datetime_format="dd-mmm-yyyy") as writer:
                wb = writer.book
                wb.set_properties({
                    "title": "MTD Dynamic Performance Dashboard",
                    "subject": "Branch performance dashboard",
                    "author": "Jaseer Pykarathodi",
                    "comments": "Generated from MTD_Dashboard_Q3.xlsx"
                })
                wb.add_vba_project(io.BytesIO(base64.b64decode(VBA_PROJECT_B64)), is_stream=True)
                # This must match the workbook code name inside the embedded VBA project;
                # otherwise Workbook_SheetChange will not fire when the dashboard date changes.
                wb.set_vba_name("ThisWorkbook1")
                ws = wb.add_worksheet("Dashboard")
                ws.set_vba_name("Sheet1")

                # Theme
                NAVY, BLUE, TEAL, GREEN = "#153B5B", "#2673C9", "#12A7A0", "#24A148"
                ORANGE, RED, BG, WHITE = "#F59E0B", "#D64545", "#F3F7FB", "#FFFFFF"
                TEXT, MUTED, BORDER = "#203040", "#64748B", "#D7E2EC"

                title_fmt = wb.add_format({"font_name":"Aptos Display","font_size":22,"bold":True,"font_color":WHITE,"bg_color":NAVY,"align":"left","valign":"vcenter"})
                subtitle_fmt = wb.add_format({"font_name":"Aptos","font_size":10,"font_color":"#DCEAF5","bg_color":NAVY,"align":"left","valign":"vcenter"})
                section_fmt = wb.add_format({"font_name":"Aptos Display","font_size":12,"bold":True,"font_color":WHITE,"bg_color":BLUE,"align":"left","valign":"vcenter"})
                label_fmt = wb.add_format({"font_name":"Aptos","font_size":9,"bold":True,"font_color":MUTED,"bg_color":WHITE,"align":"left","valign":"vcenter","border":1,"border_color":BORDER})
                selector_fmt = wb.add_format({"font_name":"Aptos","font_size":11,"bold":True,"font_color":NAVY,"bg_color":"#FFF3CD","align":"center","valign":"vcenter","border":2,"border_color":ORANGE,"locked":False})
                selector_date_fmt = wb.add_format({"font_name":"Aptos","font_size":11,"bold":True,"font_color":NAVY,"bg_color":"#FFF3CD","align":"center","valign":"vcenter","border":2,"border_color":ORANGE,"locked":False,"num_format":"dd-mmm-yyyy"})
                selector_month_fmt = wb.add_format({"font_name":"Aptos","font_size":11,"bold":True,"font_color":NAVY,"bg_color":"#FFF3CD","align":"center","valign":"vcenter","border":2,"border_color":ORANGE,"locked":False,"num_format":"mmm-yyyy"})
                card_label = wb.add_format({"font_name":"Aptos","font_size":9,"bold":True,"font_color":MUTED,"bg_color":WHITE,"align":"center","valign":"vcenter","top":1,"left":1,"right":1,"border_color":BORDER})
                card_money = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":NAVY,"bg_color":WHITE,"num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_num = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":NAVY,"bg_color":WHITE,"num_format":"#,##0;[Red](#,##0);-","align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_pct = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":NAVY,"bg_color":WHITE,"num_format":"0.0%;[Red](0.0%);-","align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_money_blue = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":"#155AA8","bg_color":"#EAF3FF","num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_money_green = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":"#13795B","bg_color":"#EAF8F3","num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_money_red = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":RED,"bg_color":"#FFF0F0","num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_num_purple = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":"#6D4CC2","bg_color":"#F2EEFF","num_format":"#,##0;[Red](#,##0);-","align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_pct_orange = wb.add_format({"font_name":"Aptos Display","font_size":17,"bold":True,"font_color":"#B56A00","bg_color":"#FFF6E5","num_format":"0.0%;[Red](0.0%);-","align":"center","valign":"vcenter","left":1,"right":1,"border_color":BORDER})
                card_foot = wb.add_format({"font_name":"Aptos","font_size":8,"font_color":MUTED,"bg_color":WHITE,"align":"center","valign":"vcenter","bottom":1,"left":1,"right":1,"border_color":BORDER})
                insight_tag = wb.add_format({"font_name":"Aptos","font_size":9,"bold":True,"font_color":WHITE,"bg_color":TEAL,"align":"center","valign":"vcenter","border":1,"border_color":TEAL})
                insight_text = wb.add_format({"font_name":"Aptos","font_size":10,"font_color":TEXT,"bg_color":WHITE,"align":"left","valign":"vcenter","border":1,"border_color":BORDER})
                small_header = wb.add_format({"font_name":"Aptos","font_size":9,"bold":True,"font_color":WHITE,"bg_color":NAVY,"align":"center","valign":"vcenter","border":1,"border_color":WHITE})
                small_date = wb.add_format({"font_name":"Aptos","font_size":8,"font_color":TEXT,"num_format":"dd-mmm","align":"center"})
                small_money = wb.add_format({"font_name":"Aptos","font_size":8,"font_color":TEXT,"num_format":"#,##0"})
                small_pct = wb.add_format({"font_name":"Aptos","font_size":8,"font_color":TEXT,"num_format":"0.0%"})

                # Controlled monthly data-entry sheet
                ws_data = wb.add_worksheet("Data Update")
                ws_data.set_vba_name("Sheet2")
                data_headers = list(data.columns) + ["Record Status"]
                input_header = wb.add_format({"font_name":"Calibri","font_size":10,"bold":True,"font_color":WHITE,"bg_color":NAVY,"align":"center","valign":"vcenter","border":1,"border_color":WHITE})
                closed_text = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":"#52606D","bg_color":"#EEF2F7","border":1,"border_color":BORDER,"locked":False})
                closed_date = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":"#52606D","bg_color":"#EEF2F7","num_format":"dd-mmm-yyyy","align":"center","border":1,"border_color":BORDER,"locked":False})
                closed_num = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":"#52606D","bg_color":"#EEF2F7","num_format":"#,##0.00;[Red](#,##0.00);-","border":1,"border_color":BORDER,"locked":False})
                helper_date = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":"#52606D","bg_color":"#EEF2F7","num_format":"dd-mmm-yyyy","align":"center","border":1,"border_color":BORDER,"locked":True})
                input_text = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":TEXT,"bg_color":"#FFF8DD","border":1,"border_color":"#E6B94A","locked":False})
                input_date = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":TEXT,"bg_color":"#FFF8DD","num_format":"dd-mmm-yyyy","align":"center","border":1,"border_color":"#E6B94A","locked":False})
                input_num = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":TEXT,"bg_color":"#FFF8DD","num_format":"#,##0.00;[Red](#,##0.00);-","border":1,"border_color":"#E6B94A","locked":False})
                formula_text = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":MUTED,"bg_color":"#F8FAFC","align":"center","border":1,"border_color":BORDER,"locked":True})
                status_fmt = wb.add_format({"font_name":"Calibri","font_size":9,"bold":True,"font_color":NAVY,"bg_color":"#F8FAFC","align":"center","border":1,"border_color":BORDER,"locked":True})
                control_title = wb.add_format({"font_name":"Calibri","font_size":12,"bold":True,"font_color":WHITE,"bg_color":TEAL,"align":"center","valign":"vcenter","border":1,"border_color":TEAL})
                control_label = wb.add_format({"font_name":"Calibri","font_size":9,"bold":True,"font_color":NAVY,"bg_color":"#EAF3FF","align":"left","valign":"vcenter","border":1,"border_color":BORDER})
                control_value_open = wb.add_format({"font_name":"Calibri","font_size":11,"bold":True,"font_color":"#116149","bg_color":"#DDF4E9","align":"center","valign":"vcenter","border":1,"border_color":"#8CD2B5"})
                control_value_closed = wb.add_format({"font_name":"Calibri","font_size":11,"bold":True,"font_color":"#B42318","bg_color":"#FDE2E2","align":"center","valign":"vcenter","border":1,"border_color":"#E7A09A"})
                control_note = wb.add_format({"font_name":"Calibri","font_size":9,"font_color":TEXT,"bg_color":WHITE,"text_wrap":True,"align":"left","valign":"top","border":1,"border_color":BORDER})

                ws_data.write_row(0, 0, data_headers, input_header)
                start_date_formula = f'DATE({open_month_start.year},{open_month_start.month},1)'
                end_date_formula = f'DATE({open_month_end.year},{open_month_end.month},{open_month_end.day})'
                entry_records = data.to_dict("records") + new_branch_template_rows
                for row_idx in range(1, DATA_LAST_ROW):
                    excel_row = row_idx + 1
                    has_source = row_idx <= len(entry_records)
                    source = entry_records[row_idx-1] if has_source else None
                    is_placeholder = bool(source.get("_new_branch_placeholder", False)) if has_source else False
                    source_date = pd.Timestamp(source["Date"]) if has_source and pd.notna(source.get("Date")) else None
                    row_is_open = bool(PERIOD_STATUS == "OPEN" and has_source and (is_placeholder or source_date.to_period("M") == latest_period))
                    blank_is_open = bool(PERIOD_STATUS == "OPEN" and not has_source)
                    editable = row_is_open or blank_is_open
                    text_fmt = input_text if editable else closed_text
                    date_fmt = input_date if editable else closed_date
                    num_fmt = input_num if editable else closed_num

                    if has_source:
                        ws_data.write(row_idx, 0, source["Branch"], text_fmt)
                        if source_date is None:
                            ws_data.write_blank(row_idx, 1, None, date_fmt)
                        else:
                            ws_data.write_datetime(row_idx, 1, source_date.to_pydatetime(), date_fmt)
                        for col_idx, column in enumerate(data.columns[3:], start=3):
                            value = source[column]
                            ws_data.write_blank(row_idx, col_idx, None, num_fmt) if pd.isna(value) else ws_data.write_number(row_idx, col_idx, float(value), num_fmt)
                        if source_date is not None:
                            ws_data.write_datetime(row_idx, 26, source_date.to_pydatetime(), helper_date)
                    else:
                        ws_data.write_blank(row_idx, 0, None, text_fmt)
                        ws_data.write_blank(row_idx, 1, None, date_fmt)
                        for col_idx in range(3, 11):
                            ws_data.write_blank(row_idx, col_idx, None, num_fmt)

                    cached_day = source.get("Day", "") if has_source else ""
                    ws_data.write_formula(row_idx, 2, f'=IF(B{excel_row}="","",TEXT(B{excel_row},"ddd"))', formula_text, cached_day)
                    if PERIOD_STATUS == "OPEN":
                        status_formula = (f'=IF(OR(A{excel_row}="",B{excel_row}=""),"",IF(COUNTIF(Entry_Branch_List,A{excel_row})=0,"INVALID",'
                                          f'IF(AND(B{excel_row}>={start_date_formula},B{excel_row}<={end_date_formula}),'
                                          f'"OPEN",IF(AND(AA{excel_row}<>"",B{excel_row}=AA{excel_row},B{excel_row}<{start_date_formula}),"CLOSED","INVALID"))))')
                        cached_status = "OPEN" if source_date is not None and source_date.to_period("M") == latest_period else ("CLOSED" if source_date is not None else "")
                    else:
                        status_formula = (f'=IF(OR(A{excel_row}="",B{excel_row}=""),"",IF(COUNTIF(Entry_Branch_List,A{excel_row})=0,"INVALID",'
                                          f'IF(AND(AA{excel_row}<>"",B{excel_row}=AA{excel_row},B{excel_row}<={end_date_formula}),"CLOSED","INVALID")))')
                        cached_status = "CLOSED" if source_date is not None else ""
                    ws_data.write_formula(row_idx, 11, status_formula, status_fmt, cached_status)

                ws_data.autofilter(0, 0, DATA_LAST_ROW-1, 11)
                ws_data.freeze_panes(1, 3)
                ws_data.hide_gridlines(2)
                ws_data.set_zoom(85)
                ws_data.set_row(0, 28)
                ws_data.set_column("A:A", 27)
                ws_data.set_column("B:B", 13, wb.add_format({"num_format":"dd-mmm-yyyy"}))
                ws_data.set_column("C:C", 9)
                ws_data.set_column("D:K", 15, wb.add_format({"num_format":"#,##0.00;[Red](#,##0.00);-"}))
                ws_data.set_column("L:L", 14)
                ws_data.set_column("M:M", 2)
                ws_data.set_column("N:Q", 16)
                ws_data.set_column("X:AA", 14, None, {"hidden":True})
                ws_data.set_tab_color(GREEN if PERIOD_STATUS == "OPEN" else RED)

                # Month-control panel. Branch setup remains controlled by the notebook without displaying instructions in Excel.
                ws_data.merge_range("N1:Q1", "MONTH CONTROL", control_title)
                ws_data.write("N2", "Controlled Month", control_label); ws_data.merge_range("O2:Q2", default_month, control_value_open)
                ws_data.write("N3", "Period Status", control_label); ws_data.merge_range("O3:Q3", PERIOD_STATUS, control_value_open if PERIOD_STATUS == "OPEN" else control_value_closed)
                ws_data.write("N4", "Allowed Dates", control_label); ws_data.merge_range("O4:Q4", f"{open_month_start:%d-%b-%Y} to {open_month_end:%d-%b-%Y}", control_value_open)

                # Hidden helper lists and immutable original dates support filters and validation.
                ws_data.write(0, 23, "Branch")
                for i, branch in enumerate(["All Branches"] + branches, start=1): ws_data.write(i, 23, branch)
                ws_data.write(0, 24, "Month")
                ws_data.write(1, 24, "All Months")
                for i, month in enumerate(months, start=2): ws_data.write_datetime(i, 24, month.to_timestamp().to_pydatetime(), wb.add_format({"num_format":"mmm-yyyy"}))
                ws_data.write(0, 25, "Date")
                ws_data.write(1, 25, "All Dates")
                for i, dt in enumerate(date_options, start=2): ws_data.write_datetime(i, 25, pd.Timestamp(dt).to_pydatetime(), wb.add_format({"num_format":"dd-mmm-yyyy"}))
                ws_data.write(0, 26, "Original Date")
                wb.define_name("Branch_List", f"='Data Update'!$X$2:$X${len(branches)+2}")
                wb.define_name("Entry_Branch_List", f"='Data Update'!$X$3:$X${len(branches)+2}")
                wb.define_name("Month_List", f"='Data Update'!$Y$2:$Y${len(months)+2}")
                wb.define_name("Date_List", f"='Data Update'!$Z$2:$Z${len(date_options)+2}")
                ws_data.data_validation(f"A2:A{DATA_LAST_ROW}", {"validate":"list","source":"=Entry_Branch_List","error_type":"stop","error_title":"Branch Not Approved","error_message":"Select an approved branch from the list. Add new branches through the notebook.","show_input":False})
                if PERIOD_STATUS == "OPEN":
                    ws_data.data_validation(f"B2:B{DATA_LAST_ROW}", {"validate":"date","criteria":"between","minimum":open_month_start.to_pydatetime(),"maximum":open_month_end.to_pydatetime(),"error_title":"Month Locked","error_message":f"Only dates from {default_month} are allowed.","show_input":False})
                ws_data.conditional_format(f"L2:L{DATA_LAST_ROW}", {"type":"text","criteria":"containing","value":"OPEN","format":wb.add_format({"font_color":"#116149","bg_color":"#DDF4E9","bold":True})})
                ws_data.conditional_format(f"L2:L{DATA_LAST_ROW}", {"type":"text","criteria":"containing","value":"CLOSED","format":wb.add_format({"font_color":"#52606D","bg_color":"#EEF2F7","bold":True})})
                ws_data.conditional_format(f"L2:L{DATA_LAST_ROW}", {"type":"text","criteria":"containing","value":"INVALID","format":wb.add_format({"font_color":"#B42318","bg_color":"#FDE2E2","bold":True})})
                # Protect formulas and the hidden branch master; all operational data-entry cells remain unlocked.
                ws_data.protect(PASSWORD, {"select_locked_cells":False,"select_unlocked_cells":True,"autofilter":True,"sort":True,"format_rows":False,"format_columns":False,"insert_rows":False,"delete_rows":False})

                # Dashboard — wide executive layout based on the supplied reference model.
                ws.hide_gridlines(2); ws.set_tab_color(TEAL); ws.set_zoom(75)
                ws.set_landscape(); ws.fit_to_pages(1, 2); ws.set_margins(0.2, 0.2, 0.3, 0.3)
                ws.set_column("A:A", 2); ws.set_column("B:V", 10.5); ws.set_column("W:W", 2)
                # Rebalance the filter strip so long branch names and labels remain fully visible.
                ws.set_column("B:B", 9); ws.set_column("C:D", 15)
                ws.set_column("E:E", 8); ws.set_column("F:G", 12)
                ws.set_column("H:H", 8); ws.set_column("I:J", 12)
                ws.set_column("K:K", 15); ws.set_column("L:V", 8.5)
                ws.set_column("X:AG", 14, None, {"hidden":True})

                ref_title = wb.add_format({"font_name":"Calibri","font_size":23,"bold":True,"font_color":WHITE,"bg_color":"#17365D","align":"left","valign":"vcenter"})
                ref_subtitle = wb.add_format({"font_name":"Calibri","font_size":10,"font_color":"#DCEBFF","bg_color":"#17365D","align":"left","valign":"vcenter"})
                ref_exec = wb.add_format({"font_name":"Calibri","font_size":10,"bold":True,"font_color":WHITE,"bg_color":TEAL,"align":"center","valign":"vcenter"})
                ref_section = wb.add_format({"font_name":"Calibri","font_size":11,"bold":True,"font_color":WHITE,"bg_color":TEAL,"align":"left","valign":"vcenter"})
                ref_filter_label = wb.add_format({"font_name":"Calibri","font_size":11,"bold":True,"font_color":"#102A43","bg_color":"#EAF3FF","align":"center","valign":"vcenter","border":1,"border_color":"#C7D9EA"})
                ref_available = wb.add_format({"font_name":"Calibri","font_size":10,"bold":True,"font_color":"#17365D","bg_color":"#F4F7FB","align":"center","valign":"vcenter","border":1,"border_color":"#C7D9EA"})
                applied_month_fmt = wb.add_format({"font_name":"Calibri","font_size":11,"bold":True,"font_color":"#0F766E","bg_color":"#EAF8F3","num_format":"mmm-yyyy","align":"center","valign":"vcenter","border":1,"border_color":"#C7D9EA"})
                kpi_head = wb.add_format({"font_name":"Calibri","font_size":10,"bold":True,"font_color":"#17365D","bg_color":"#EAF3FF","align":"center","valign":"vcenter","border":1,"border_color":"#D5E2EF"})
                kpi_money = wb.add_format({"font_name":"Calibri","font_size":18,"bold":True,"font_color":"#2563EB","bg_color":"#EAF3FF","num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","border":1,"border_color":"#D5E2EF"})
                kpi_green = wb.add_format({"font_name":"Calibri","font_size":18,"bold":True,"font_color":"#0F9D8A","bg_color":"#EAF8F3","num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","border":1,"border_color":"#D5E2EF"})
                kpi_number = wb.add_format({"font_name":"Calibri","font_size":18,"bold":True,"font_color":"#6D4CC2","bg_color":"#F2EEFF","num_format":"#,##0;[Red](#,##0);-","align":"center","valign":"vcenter","border":1,"border_color":"#D5E2EF"})
                kpi_pct = wb.add_format({"font_name":"Calibri","font_size":18,"bold":True,"font_color":"#B56A00","bg_color":"#FFF6E5","num_format":"0.0%;[Red](0.0%);-","align":"center","valign":"vcenter","border":1,"border_color":"#D5E2EF"})
                kpi_var = wb.add_format({"font_name":"Calibri","font_size":18,"bold":True,"font_color":RED,"bg_color":"#FFF0F0","num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","border":1,"border_color":"#D5E2EF"})
                kpi_text = wb.add_format({"font_name":"Calibri","font_size":13,"bold":True,"font_color":"#17365D","bg_color":"#EAF8F3","align":"center","valign":"vcenter","text_wrap":True,"border":1,"border_color":"#D5E2EF"})
                kpi_change = wb.add_format({"font_name":"Calibri","font_size":17,"bold":True,"font_color":"#B56A00","bg_color":"#FFF6E5","align":"center","valign":"vcenter","border":1,"border_color":"#D5E2EF"})
                forecast_note = wb.add_format({"font_name":"Calibri","font_size":9,"italic":True,"font_color":"#52606D","bg_color":"#F4F7FB","align":"left","valign":"vcenter","border":1,"border_color":"#D5E2EF"})

                ws.merge_range("B2:V3", "MTD PERFORMANCE DASHBOARD", ref_title)
                ws.merge_range("B4:Q4", "Branch performance, targets, bills, average bill value and management comparisons", ref_subtitle)
                ws.merge_range("R4:V4", f"EXECUTIVE VIEW • {default_month.upper()} {PERIOD_STATUS}", ref_exec)
                ws.merge_range("B6:V6", "REPORT FILTERS", ref_section)
                ws.set_row(6, 23)
                ws.write("B7", "Branch", ref_filter_label); ws.merge_range("C7:D7", default_branch, selector_fmt)
                ws.data_validation("C7:D7", {"validate":"list","source":"=Branch_List","show_input":False})
                ws.write("E7", "Month", ref_filter_label); ws.merge_range("F7:G7", "", selector_month_fmt)
                ws.write_datetime("F7", latest_period.to_timestamp().to_pydatetime(), selector_month_fmt)
                ws.data_validation("F7:G7", {"validate":"list","source":"=Month_List","show_input":False})
                ws.write("H7", "Date", ref_filter_label); ws.merge_range("I7:J7", "All Dates", selector_date_fmt)
                ws.data_validation("I7:J7", {"validate":"list","source":"=Date_List","show_input":False})
                # A specific date overrides the month filter without relying on VBA. Show the
                # month selector as inactive and display the effective month beside it.
                month_inactive_fmt = wb.add_format({"bg_color":"#EEF2F7","font_color":"#52606D","bold":True,"num_format":'"DATE MODE"',"border":1,"border_color":"#C7D9EA","align":"center","valign":"vcenter"})
                ws.conditional_format("F7:G7", {"type":"formula","criteria":'=ISNUMBER($I$7)',"format":month_inactive_fmt})
                ws.write("K7", "Applied Month", ref_filter_label)
                ws.merge_range("L7:M7", "", applied_month_fmt)
                applied_month_cached = int((latest_period.to_timestamp() - pd.Timestamp("1899-12-30")).days)
                ws.write_formula("L7", '=IF(ISNUMBER($I$7),DATE(YEAR($I$7),MONTH($I$7),1),$F$7)', applied_month_fmt, applied_month_cached)
                available_text = f"{min(dates):%d-%b-%Y} to {max(dates):%d-%b-%Y}"
                applied_default = f"Applied: {latest_period.to_timestamp():%b-%Y} • All Dates | Available: {available_text}"
                applied_formula = (f'=IF(ISNUMBER($I$7),"Applied: "&TEXT($I$7,"dd-mmm-yyyy")&" ("&TEXT($I$7,"mmm-yyyy")&")",'
                                   f'IF(ISNUMBER($F$7),"Applied: "&TEXT($F$7,"mmm-yyyy")&" • All Dates","Applied: All Available Dates"))'
                                   f'&" | Available: {available_text}"')
                ws.merge_range("N7:V7", "", ref_available)
                ws.write_formula("N7", applied_formula, ref_available, applied_default)
                ws.merge_range("B9:V9", "KEY PERFORMANCE INDICATORS", ref_section)

                def sumifs_formula(col):
                    value=f"'Data Update'!${col}$2:${col}$5000"; branch="'Data Update'!$A$2:$A$5000"; dr="'Data Update'!$B$2:$B$5000"; status="'Data Update'!$L$2:$L$5000"
                    return (f'=IF(ISNUMBER($I$7),IF($C$7="All Branches",SUMIFS({value},{dr},$I$7,{status},"<>INVALID"),SUMIFS({value},{branch},$C$7,{dr},$I$7,{status},"<>INVALID")),'
                            f'IF(NOT(ISNUMBER($F$7)),IF($C$7="All Branches",SUMIFS({value},{status},"<>INVALID"),SUMIFS({value},{branch},$C$7,{status},"<>INVALID")),'
                            f'IF($C$7="All Branches",SUMIFS({value},{dr},">="&$F$7,{dr},"<="&EOMONTH($F$7,0),{status},"<>INVALID"),SUMIFS({value},{branch},$C$7,{dr},">="&$F$7,{dr},"<="&EOMONTH($F$7,0),{status},"<>INVALID"))))')
                def previous_sales_formula():
                    value="'Data Update'!$E$2:$E$5000"; branch="'Data Update'!$A$2:$A$5000"; dr="'Data Update'!$B$2:$B$5000"; status="'Data Update'!$L$2:$L$5000"
                    return (f'=IF(OR($X$4=0,$X$5=0),0,IF($C$7="All Branches",'
                            f'SUMIFS({value},{dr},">="&$X$4,{dr},"<="&$X$5,{status},"<>INVALID"),'
                            f'SUMIFS({value},{branch},$C$7,{dr},">="&$X$4,{dr},"<="&$X$5,{status},"<>INVALID")))')
                def month_to_cutoff_formula(col):
                    value=f"'Data Update'!${col}$2:${col}$5000"; branch="'Data Update'!$A$2:$A$5000"; dr="'Data Update'!$B$2:$B$5000"; status="'Data Update'!$L$2:$L$5000"
                    return (f'=IF(OR($X$6=0,$X$7=0),0,IF($C$7="All Branches",'
                            f'SUMIFS({value},{dr},">="&$X$6,{dr},"<="&$X$7,{status},"<>INVALID"),'
                            f'SUMIFS({value},{branch},$C$7,{dr},">="&$X$6,{dr},"<="&$X$7,{status},"<>INVALID")))')
                helper_start=82
                branch_first_excel=helper_start+2
                branch_last_excel=helper_start+1+len(branches)
                card_blocks=[("B","D"),("E","G"),("H","J"),("K","M"),("N","Q"),("R","V")]
                kpis1=[("◎ SALES TARGET",sumifs_formula("D"),sales_target,kpi_money),("◆ SALES ACHIEVED",sumifs_formula("E"),sales_actual,kpi_green),("● SALES ACHIEVEMENT",'=IFERROR(E11/B11,0)',sales_pct,kpi_pct),("▲▼ SALES VARIANCE",'=E11-B11',variance,kpi_var),("◎ NOB TARGET",sumifs_formula("F"),nob_target,kpi_number),("◆ NOB ACHIEVED",sumifs_formula("G"),nob_actual,kpi_number)]
                previous_display = float(previous_sales) if previous_sales else "Not available"
                change_display = (("▲ " if period_change >= 0 else "▼ ") + f"{abs(period_change):.1%}") if previous_sales else "Not available"
                top_branch_formula=(f'=INDEX($AC${branch_first_excel}:$AC${branch_last_excel},'
                                    f'MATCH(MAX($AF${branch_first_excel}:$AF${branch_last_excel}),'
                                    f'$AF${branch_first_excel}:$AF${branch_last_excel},0))')
                kpis2=[("● NOB ACHIEVEMENT",'=IFERROR(R11/N11,0)',nob_pct,kpi_pct),("◎ ABV TARGET",'=IFERROR(B11/N11,0)',abv_target,kpi_money),("◆ ABV ACTUAL",'=IFERROR(E11/R11,0)',abv_actual,kpi_green),("◀ PREVIOUS PERIOD SALES",'=IF(X2=0,"Not available",X2)',previous_display,kpi_money),("▲▼ PERIOD CHANGE",'=IF(X2=0,"Not available",IF(X3>=0,"▲ ","▼ ")&TEXT(ABS(X3),"0.0%"))',change_display,kpi_change),("★ TOP BRANCH",top_branch_formula,best_default,kpi_text)]
                for row_label,row_value,kpis in [(10,11,kpis1),(15,16,kpis2)]:
                    for (c1,c2),(label,formula,cached,fmt) in zip(card_blocks,kpis):
                        ws.merge_range(f"{c1}{row_label}:{c2}{row_label}",label,kpi_head)
                        ws.merge_range(f"{c1}{row_value}:{c2}{row_value+2}","",fmt)
                        ws.write_formula(f"{c1}{row_value}",formula,fmt,cached)
                latest_data_date = pd.Timestamp(max(dates))
                latest_month_start = latest_period.to_timestamp()
                controlled_dates = data.loc[data["Date"].dt.to_period("M") == latest_period, "Date"]
                latest_controlled_date = pd.Timestamp(controlled_dates.max()) if len(controlled_dates) else None
                latest_month_serial = int((latest_month_start - pd.Timestamp("1899-12-30")).days)
                previous_start_serial = int((previous_period.to_timestamp() - pd.Timestamp("1899-12-30")).days)
                previous_end_serial = int((min(previous_cutoff, previous_month_end) - pd.Timestamp("1899-12-30")).days)
                latest_date_serial = int((latest_controlled_date - pd.Timestamp("1899-12-30")).days) if latest_controlled_date is not None else 0
                previous_start_formula = '=IF(ISNUMBER($I$7),$I$7-1,IF(ISNUMBER($F$7),$F$7-DAY($F$7-1),0))'
                controlled_latest_day = latest_controlled_date.day if latest_controlled_date is not None else 0
                elapsed_day_formula = f'IF($F$7={latest_month_serial},{controlled_latest_day},$F$7+32-DAY($F$7+32)+1-$F$7)'
                previous_end_formula = f'=IF(ISNUMBER($I$7),$I$7-1,IF(ISNUMBER($F$7),$X$4+MIN({elapsed_day_formula},$F$7-$X$4)-1,0))'
                ws.write_formula("X4",previous_start_formula,small_date,previous_start_serial)
                ws.write_formula("X5",previous_end_formula,small_date,previous_end_serial)
                ws.write_formula("X2",previous_sales_formula(),kpi_money,float(previous_sales))
                ws.write_formula("X3",'=IFERROR((E11-X2)/X2,0)',kpi_pct,float(period_change))
                ws.write_formula("X6",'=IF(ISNUMBER($I$7),DATE(YEAR($I$7),MONTH($I$7),1),IF(ISNUMBER($F$7),$F$7,0))',small_date,latest_month_serial)
                ws.write_formula("X7",'=IF($X$6=0,0,IF(ISNUMBER($I$7),IF(COUNTIFS(\'Data Update\'!$B$2:$B$5000,$I$7,\'Data Update\'!$L$2:$L$5000,"<>INVALID")>0,$I$7,0),IFERROR(LOOKUP(2,1/((\'Data Update\'!$B$2:$B$5000>=$X$6)*(\'Data Update\'!$B$2:$B$5000<=EOMONTH($X$6,0))*(\'Data Update\'!$L$2:$L$5000<>"INVALID")),\'Data Update\'!$B$2:$B$5000),0)))',small_date,latest_date_serial)
                ws.write_formula("X8",'=IF(X7=0,0,DAY(X7))',small_money,float(elapsed_days))
                ws.write_formula("X9",'=IF(X6=0,0,DAY(EOMONTH(X6,0)))',small_money,float(days_in_month))
                ws.write_formula("X10",'=MAX(X9-X8,0)',small_money,float(days_remaining))
                ws.write_formula("X11",month_to_cutoff_formula("E"),small_money,float(sales_actual))
                ws.write_formula("X12",month_to_cutoff_formula("D"),small_money,float(sales_target))
                ws.write_formula("X13",'=IFERROR(X11/X8*X9,0)',small_money,float(month_end_forecast))
                ws.write_formula("X14",'=IFERROR(X12/X8*X9,0)',small_money,float(projected_month_target))
                ws.write_formula("X15",'=X13-X14',small_money,float(forecast_gap))
                ws.write_formula("X16",'=MAX(X14-X11,0)',small_money,float(remaining_target))
                ws.write_formula("X17",'=IF(X10=0,0,X16/X10)',small_money,float(required_daily_sales))
                ws.conditional_format("H11:J13",{"type":"cell","criteria":">=","value":1,"format":wb.add_format({"font_color":GREEN,"bold":True})})
                ws.conditional_format("H11:J13",{"type":"cell","criteria":"<","value":1,"format":wb.add_format({"font_color":RED,"bold":True})})
                ws.conditional_format("B16:D18",{"type":"cell","criteria":">=","value":1,"format":wb.add_format({"font_color":GREEN,"bold":True})})
                ws.conditional_format("N16:Q18",{"type":"formula","criteria":'=$X$3>=0',"format":wb.add_format({"font_color":GREEN,"bold":True})})
                ws.conditional_format("N16:Q18",{"type":"formula","criteria":'=$X$3<0',"format":wb.add_format({"font_color":RED,"bold":True})})

                ws.merge_range("B20:V20","DAILY EXCEPTION REPORT",ref_section)
                tags=[]; texts=[]
                for helper_row in range(branch_first_excel, branch_last_excel + 1):
                    visible=f'OR($C$7="All Branches",$C$7=$AC${helper_row})'
                    tags.append(f'=IF({visible},IF(AND($AD${helper_row}=0,$AE${helper_row}=0),"ℹ DATA",IF($AF${helper_row}>=1,"▲ POSITIVE",IF($AF${helper_row}>=0.85,"● WATCH","⚠ CRITICAL"))),"")')
                    texts.append(f'=IF({visible},IF(AND($AD${helper_row}=0,$AE${helper_row}=0),$AC${helper_row}&" — No data available for the selected date or period.",$AC${helper_row}&" — Sales "&TEXT($AF${helper_row},"0.0%")&"; "&IF($AE${helper_row}>=$AD${helper_row},"target exceeded by SAR ","target gap SAR ")&TEXT(ABS($AE${helper_row}-$AD${helper_row}),"#,##0")&"; bills "&TEXT($AG${helper_row},"0.0%")&". "&IF($AF${helper_row}>=1,"Maintain the current pace.",IF($AF${helper_row}>=0.85,"Close the remaining gap and review ABV.","Start a daily recovery plan and review conversion."))),"")')
                while len(tags) < 6:
                    tags.append('=""'); texts.append('=""')
                for i in range(6):
                    row=21+i; tag,msg=insights[i] if i<len(insights) else ("","")
                    ws.merge_range(row-1,1,row-1,3,"",insight_tag); ws.write_formula(row-1,1,tags[i],insight_tag,tag)
                    ws.merge_range(row-1,4,row-1,21,"",insight_text); ws.write_formula(row-1,4,texts[i],insight_text,msg); ws.set_row(row-1,23)

                ws.merge_range("B28:V28","FORECAST & PACE",ref_section)
                forecast_available = latest_controlled_date is not None
                forecast_kpis=[
                    ("◆ MONTH-END FORECAST",'=IF(OR($X$6=0,$X$7=0),"Not available",$X$13)',month_end_forecast if forecast_available else "Not available",kpi_green),
                    ("◎ PROJECTED MONTH TARGET",'=IF(OR($X$6=0,$X$7=0),"Not available",$X$14)',projected_month_target if forecast_available else "Not available",kpi_money),
                    ("▲▼ FORECAST GAP",'=IF(OR($X$6=0,$X$7=0),"Not available",$X$15)',forecast_gap if forecast_available else "Not available",kpi_var),
                    ("◎ REMAINING TO TARGET",'=IF(OR($X$6=0,$X$7=0),"Not available",$X$16)',remaining_target if forecast_available else "Not available",kpi_money),
                    ("◆ REQUIRED DAILY SALES",'=IF(OR($X$6=0,$X$7=0),"Not available",IF($X$10=0,"Period closed",$X$17))',("Not available" if not forecast_available else ("Period closed" if days_remaining == 0 else required_daily_sales)),kpi_green),
                    ("◷ DAYS REMAINING",'=IF(OR($X$6=0,$X$7=0),"Not available",IF($X$10=0,"Closed",$X$10))',("Not available" if not forecast_available else ("Closed" if days_remaining == 0 else days_remaining)),kpi_number),
                ]
                for (c1,c2),(label,formula,cached,fmt) in zip(card_blocks,forecast_kpis):
                    ws.merge_range(f"{c1}29:{c2}29",label,kpi_head)
                    ws.merge_range(f"{c1}30:{c2}32","",fmt)
                    ws.write_formula(f"{c1}30",formula,fmt,cached)
                ws.conditional_format("H30:J32",{"type":"formula","criteria":'=$X$15>=0',"format":wb.add_format({"font_color":GREEN,"bold":True})})
                ws.conditional_format("H30:J32",{"type":"formula","criteria":'=$X$15<0',"format":wb.add_format({"font_color":RED,"bold":True})})
                ws.merge_range("B33:V33","",forecast_note)
                ws.set_row(32,18)

                start=helper_start
                ws.write_row(start,23,["Date","Sales Target","Sales Achieved","NOB Target","NOB Achieved"],small_header)
                for idx,dt in enumerate(dates,start=start+1):
                    d=pd.Timestamp(dt); day=data[data["Date"]==d]; ws.write_datetime(idx,23,d.to_pydatetime(),small_date)
                    for j,col in enumerate(["D","E","F","G"],start=24):
                        source=["Sales Target","Sales Achieved","NOB Target","NOB Achieved"][j-24]; cached=day[source].sum(min_count=1); show=d.strftime("%b-%Y")==default_month
                        formula=f'=IF(AND(OR(NOT(ISNUMBER($F$7)),EOMONTH(X{idx+1},0)=EOMONTH($F$7,0)),OR(NOT(ISNUMBER($I$7)),X{idx+1}=$I$7)),IF($C$7="All Branches",SUMIFS(\'Data Update\'!${col}$2:${col}$5000,\'Data Update\'!$B$2:$B$5000,X{idx+1},\'Data Update\'!$L$2:$L$5000,"<>INVALID"),SUMIFS(\'Data Update\'!${col}$2:${col}$5000,\'Data Update\'!$A$2:$A$5000,$C$7,\'Data Update\'!$B$2:$B$5000,X{idx+1},\'Data Update\'!$L$2:$L$5000,"<>INVALID")),NA())'
                        ws.write_formula(idx,j,formula,small_money,float(cached) if show and pd.notna(cached) else "")
                branch_start=start; ws.write_row(branch_start,28,["Branch","Target","Actual","Achievement","NOB Achievement"],small_header)
                cb=(current.groupby("Branch",as_index=False)
                    .agg({"Sales Target":"sum","Sales Achieved":"sum","NOB Target":"sum","NOB Achieved":"sum"})
                    .set_index("Branch").reindex(branches).fillna(0))
                def bformula(col,ref):
                    v=f"'Data Update'!${col}$2:${col}$5000"; br="'Data Update'!$A$2:$A$5000"; dr="'Data Update'!$B$2:$B$5000"; status="'Data Update'!$L$2:$L$5000"
                    return f'=IF(ISNUMBER($I$7),SUMIFS({v},{br},{ref},{dr},$I$7,{status},"<>INVALID"),IF(NOT(ISNUMBER($F$7)),SUMIFS({v},{br},{ref},{status},"<>INVALID"),SUMIFS({v},{br},{ref},{dr},">="&$F$7,{dr},"<="&EOMONTH($F$7,0),{status},"<>INVALID")))'
                for idx,branch in enumerate(branches,start=branch_start+1):
                    er=idx+1; vals=cb.loc[branch]; target=vals["Sales Target"]; actual=vals["Sales Achieved"]; nt=vals["NOB Target"]; na=vals["NOB Achieved"]
                    ws.write(idx,28,branch); ws.write_formula(idx,29,bformula("D",f'AC{er}'),small_money,float(target)); ws.write_formula(idx,30,bformula("E",f'AC{er}'),small_money,float(actual))
                    ws.write_formula(idx,31,f'=IFERROR(AE{er}/AD{er},0)',small_pct,float(actual/target if target else 0)); ws.write_formula(idx,32,f'=IFERROR({bformula("G",f"AC{er}")[1:]}/{bformula("F",f"AC{er}")[1:]},0)',small_pct,float(na/nt if nt else 0))
                ws.write("X64","Previous Period"); ws.write("X65","Selected Period"); ws.write_formula("Y64",'=IF(X2=0,NA(),X2)',small_money,float(previous_sales)); ws.write_formula("Y65",'=E11',small_money,float(sales_actual))

                chart_band = wb.add_format({"font_name":"Calibri","font_size":11,"bold":True,"font_color":WHITE,"bg_color":"#17365D","align":"left","valign":"vcenter","left":1,"right":1,"border_color":"#17365D"})
                ws.merge_range("B36:K36","DAILY SALES TREND • SAR",chart_band)
                ws.merge_range("L36:V36","BRANCH TARGET VS ACHIEVED • SAR",chart_band)
                ws.merge_range("B55:H55","CURRENT VS PREVIOUS PERIOD",chart_band)
                ws.merge_range("I55:P55","BRANCH PERFORMANCE HEATMAP",chart_band)
                ws.merge_range("Q55:V55","BILL ACHIEVEMENT BY BRANCH",chart_band)

                def style_powerbi_chart(chart, legend=False):
                    chart.set_title({"none":True})
                    chart.set_chartarea({"fill":{"color":"#FFFFFF"},"border":{"color":"#D7E2EC","width":1}})
                    chart.set_plotarea({"fill":{"color":"#FFFFFF"},"border":{"none":True}})
                    if legend:
                        chart.set_legend({"position":"bottom","font":{"name":"Calibri","size":9,"color":"#52606D"}})
                    else:
                        chart.set_legend({"none":True})
                    chart.show_hidden_data()

                chart1=wb.add_chart({"type":"line"})
                chart1.add_series({"name":"Daily Target","categories":["Dashboard",start+1,23,start+len(dates),23],"values":["Dashboard",start+1,24,start+len(dates),24],"line":{"color":ORANGE,"width":2,"dash_type":"dash"}})
                chart1.add_series({"name":"Daily Sales","categories":["Dashboard",start+1,23,start+len(dates),23],"values":["Dashboard",start+1,25,start+len(dates),25],"line":{"color":BLUE,"width":3}})
                chart1.set_y_axis({"num_format":"#,##0","major_gridlines":{"visible":True,"line":{"color":"#E8EEF4"}},"num_font":{"name":"Calibri","size":8,"color":"#64748B"}})
                chart1.set_x_axis({"date_axis":True,"num_format":"dd-mmm","num_font":{"name":"Calibri","size":8,"color":"#64748B"},"line":{"color":"#CBD5E1"}})
                style_powerbi_chart(chart1,True); ws.insert_chart("B37",chart1,{"x_scale":1.25,"y_scale":1.08})

                chart2=wb.add_chart({"type":"column"})
                chart2.add_series({"name":"Target","categories":["Dashboard",branch_start+1,28,branch_start+len(branches),28],"values":["Dashboard",branch_start+1,29,branch_start+len(branches),29],"fill":{"color":"#9AAAC0"},"border":{"none":True},"gap":85})
                chart2.add_series({"name":"Achieved","categories":["Dashboard",branch_start+1,28,branch_start+len(branches),28],"values":["Dashboard",branch_start+1,30,branch_start+len(branches),30],"fill":{"color":TEAL},"border":{"none":True},"data_labels":{"value":True,"num_format":"#,##0","position":"outside_end","font":{"size":8,"color":"#334155"}}})
                chart2.set_y_axis({"num_format":"#,##0","major_gridlines":{"visible":True,"line":{"color":"#E8EEF4"}},"num_font":{"name":"Calibri","size":8,"color":"#64748B"}})
                chart2.set_x_axis({"num_font":{"name":"Calibri","size":8,"color":"#64748B"},"line":{"color":"#CBD5E1"}})
                style_powerbi_chart(chart2,True); ws.insert_chart("L37",chart2,{"x_scale":1.38,"y_scale":1.08})

                chart3=wb.add_chart({"type":"column"})
                chart3.add_series({"name":"Sales","categories":"=Dashboard!$X$64:$X$65","values":"=Dashboard!$Y$64:$Y$65","points":[{"fill":{"color":"#9AAAC0"},"border":{"none":True}},{"fill":{"color":BLUE},"border":{"none":True}}],"data_labels":{"value":True,"num_format":"#,##0","position":"outside_end","font":{"size":9,"bold":True,"color":"#334155"}},"gap":65})
                chart3.set_y_axis({"num_format":"#,##0","major_gridlines":{"visible":True,"line":{"color":"#EEF2F7"}},"num_font":{"name":"Calibri","size":8,"color":"#64748B"}})
                chart3.set_x_axis({"num_font":{"name":"Calibri","size":8,"color":"#64748B"},"line":{"color":"#CBD5E1"}})
                style_powerbi_chart(chart3); ws.insert_chart("B56",chart3,{"x_scale":0.9,"y_scale":0.98})

                heatmap_header = wb.add_format({"font_name":"Calibri","font_size":9,"bold":True,"font_color":WHITE,"bg_color":"#2673C9","align":"center","valign":"vcenter","border":1,"border_color":WHITE})
                heatmap_branch = wb.add_format({"font_name":"Calibri","font_size":9,"bold":True,"font_color":"#17365D","bg_color":"#F4F7FB","align":"left","valign":"vcenter","border":1,"border_color":"#D7E2EC"})
                heatmap_pct = wb.add_format({"font_name":"Calibri","font_size":11,"bold":True,"font_color":"#17365D","bg_color":WHITE,"num_format":"0.0%","align":"center","valign":"vcenter","border":1,"border_color":"#D7E2EC"})
                heatmap_gap = wb.add_format({"font_name":"Calibri","font_size":10,"bold":True,"font_color":"#17365D","bg_color":WHITE,"num_format":'\"SAR\" #,##0;[Red](\"SAR\" #,##0);-',"align":"center","valign":"vcenter","border":1,"border_color":"#D7E2EC"})
                heatmap_status = wb.add_format({"font_name":"Calibri","font_size":8,"bold":True,"font_color":"#17365D","bg_color":"#F4F7FB","align":"center","valign":"vcenter","border":1,"border_color":"#D7E2EC"})
                heat_green = wb.add_format({"font_color":"#116149","bg_color":"#DDF4E9","bold":True})
                heat_amber = wb.add_format({"font_color":"#8A5300","bg_color":"#FFF1CC","bold":True})
                heat_red = wb.add_format({"font_color":"#B42318","bg_color":"#FDE2E2","bold":True})
                ws.merge_range("I56:K56","BRANCH",heatmap_header)
                ws.merge_range("L56:M56","SALES %",heatmap_header)
                ws.merge_range("N56:O56","GAP • SAR",heatmap_header)
                ws.write("P56","STATUS",heatmap_header)
                for offset,branch in enumerate(branches):
                    display_row=57+offset; helper_row=branch_first_excel+offset
                    vals=cb.loc[branch]; target=float(vals["Sales Target"]); actual=float(vals["Sales Achieved"])
                    pct=float(actual/target if target else 0); gap=float(actual-target)
                    status="ON TRACK" if pct>=1 else ("WATCH" if pct>=0.85 else "CRITICAL")
                    ws.merge_range(f"I{display_row}:K{display_row}","",heatmap_branch); ws.write_formula(f"I{display_row}",f'=AC{helper_row}',heatmap_branch,branch)
                    ws.merge_range(f"L{display_row}:M{display_row}","",heatmap_pct); ws.write_formula(f"L{display_row}",f'=AF{helper_row}',heatmap_pct,pct)
                    ws.merge_range(f"N{display_row}:O{display_row}","",heatmap_gap); ws.write_formula(f"N{display_row}",f'=AE{helper_row}-AD{helper_row}',heatmap_gap,gap)
                    ws.write_formula(f"P{display_row}",f'=IF(L{display_row}>=1,"ON TRACK",IF(L{display_row}>=0.85,"WATCH","CRITICAL"))',heatmap_status,status)
                    ws.conditional_format(f"L{display_row}:M{display_row}",{"type":"formula","criteria":f'=$L{display_row}>=1',"format":heat_green})
                    ws.conditional_format(f"L{display_row}:M{display_row}",{"type":"formula","criteria":f'=AND($L{display_row}>=0.85,$L{display_row}<1)',"format":heat_amber})
                    ws.conditional_format(f"L{display_row}:M{display_row}",{"type":"formula","criteria":f'=$L{display_row}<0.85',"format":heat_red})
                    ws.conditional_format(f"N{display_row}:O{display_row}",{"type":"formula","criteria":f'=$N{display_row}>=0',"format":heat_green})
                    ws.conditional_format(f"N{display_row}:O{display_row}",{"type":"formula","criteria":f'=$N{display_row}<0',"format":heat_red})
                    ws.conditional_format(f"P{display_row}",{"type":"text","criteria":"containing","value":"ON TRACK","format":heat_green})
                    ws.conditional_format(f"P{display_row}",{"type":"text","criteria":"containing","value":"WATCH","format":heat_amber})
                    ws.conditional_format(f"P{display_row}",{"type":"text","criteria":"containing","value":"CRITICAL","format":heat_red})
                    ws.set_row(display_row-1,28)

                chart5=wb.add_chart({"type":"bar"})
                chart5.add_series({"name":"NOB Achievement","categories":["Dashboard",branch_start+1,28,branch_start+len(branches),28],"values":["Dashboard",branch_start+1,32,branch_start+len(branches),32],"fill":{"color":ORANGE},"border":{"none":True},"data_labels":{"value":True,"num_format":"0%","position":"outside_end","font":{"size":9,"bold":True,"color":"#334155"}},"gap":55})
                chart5.set_x_axis({"num_format":"0%","min":0,"max":1.4,"major_unit":0.2,"major_gridlines":{"visible":True,"line":{"color":"#EEF2F7"}},"num_font":{"name":"Calibri","size":8,"color":"#64748B"}})
                chart5.set_y_axis({"num_font":{"name":"Calibri","size":8,"color":"#64748B"},"line":{"none":True}})
                style_powerbi_chart(chart5); ws.insert_chart("Q56",chart5,{"x_scale":0.95,"y_scale":0.98})

                footer_fmt = wb.add_format({"font_name":"Calibri","font_size":10,"bold":True,"font_color":WHITE,"bg_color":NAVY,"align":"center","valign":"vcenter","border":1,"border_color":NAVY})
                ws.merge_range("B69:V69", "Powered by Jaseer PT | Data Analyst", footer_fmt)
                ws.set_row(68, 22)

                for hidden_row in range(71, start + len(dates) + 3):
                    ws.set_row(hidden_row, None, None, {"hidden":True})
                ws.freeze_panes(6,1); ws.set_selection("C7"); ws.set_top_left_cell("A1"); ws.activate(); ws.protect(PASSWORD,{"select_locked_cells":False,"select_unlocked_cells":True,"format_cells":False,"format_columns":False,"format_rows":False,"insert_rows":False,"delete_rows":False,"sort":False,"autofilter":False})

                # Dashboard remains the first sheet; Data Update stays visible and unprotected for updates.
                ws.set_first_sheet()
                ws.activate()

            Path(OUTPUT_FILE).write_bytes(excel_output.getvalue())
            protect_workbook_structure(OUTPUT_FILE)
            print(f"Created: {OUTPUT_FILE}")


        summary = logs.getvalue().strip()
        message = f"Dashboard created successfully: {OPEN_MONTH} | {PERIOD_STATUS}"
        if ADD_BRANCH == "YES":
            message += f" | New branch: {REQUESTED_NEW_BRANCH}"
        if summary:
            message += "\n\n" + summary[-2500:]
        return OUTPUT_FILE, message
    except Exception as exc:
        detail = logs.getvalue().strip()
        if detail:
            raise gr.Error(f"{exc}\n\n{detail[-1500:]}")
        raise gr.Error(str(exc))



# -----------------------------------------------------------------------------
# SECURE WEB APP LAYER
# -----------------------------------------------------------------------------
BUILD_VERSION = "LIVE-WEB-2026.08.25-v5"


def _credentials():
    """Read mandatory Admin/User credentials from Render environment variables."""
    admin_user = os.getenv("ADMIN_USERNAME", "").strip()
    admin_pass = os.getenv("ADMIN_PASSWORD", "").strip()
    user_user = os.getenv("USER_USERNAME", "").strip()
    user_pass = os.getenv("USER_PASSWORD", "").strip()
    return admin_user, admin_pass, user_user, user_pass


def _validate_credentials_config():
    admin_user, admin_pass, user_user, user_pass = _credentials()
    missing = []
    if not admin_user:
        missing.append("ADMIN_USERNAME")
    if not admin_pass:
        missing.append("ADMIN_PASSWORD")
    if not user_user:
        missing.append("USER_USERNAME")
    if not user_pass:
        missing.append("USER_PASSWORD")
    if missing:
        raise RuntimeError(
            "Mandatory login is enabled. Add these Render Environment Variables: "
            + ", ".join(missing)
        )


def _auth_user(username, password):
    admin_user, admin_pass, user_user, user_pass = _credentials()
    return (
        (username == admin_user and password == admin_pass)
        or (username == user_user and password == user_pass)
    )


def _role_from_request(request: gr.Request):
    admin_user, _, user_user, _ = _credentials()
    username = (getattr(request, "username", None) or "").strip()
    if username == admin_user:
        return "ADMIN", username
    if username == user_user:
        return "USER", username
    return "USER", username or "User"


def _read_control_settings(input_file):
    """Read Admin-controlled month/status saved in a previously generated workbook."""
    if not input_file:
        return None, None
    try:
        wb = load_workbook(input_file, read_only=True, data_only=False, keep_vba=True)
        if "Data Update" not in wb.sheetnames:
            wb.close()
            return None, None
        ws = wb["Data Update"]
        month = ws["O2"].value
        status = ws["O3"].value
        wb.close()
        month = str(month).strip() if month is not None else None
        status = str(status).strip().upper() if status is not None else None
        if status not in {"OPEN", "CLOSED"}:
            status = None
        if month:
            try:
                month = pd.to_datetime(month, errors="raise").strftime("%b-%Y")
            except Exception:
                # Existing generator stores Mon-YYYY as text, so keep valid text as-is.
                if not re.fullmatch(r"[A-Za-z]{3}-\d{4}", month):
                    month = None
        return month, status
    except Exception:
        return None, None


def _standardize_web_source(raw, fixed_branch=None):
    raw = raw.copy()
    raw.columns = [clean_header_for_ui(c) for c in raw.columns]
    found = {
        key: next((c for c in choices if c in raw.columns), None)
        for key, choices in ALIASES.items()
    }
    if not found["date"] or not found["sales_achieved"]:
        return None
    if fixed_branch is None and not found["branch"]:
        return None

    if fixed_branch is None:
        branch_values = raw[found["branch"]].astype(str).str.strip()
    else:
        branch_values = pd.Series([str(fixed_branch).strip()] * len(raw), index=raw.index)

    df = pd.DataFrame({
        "Branch": branch_values,
        "Date": pd.to_datetime(raw[found["date"]], errors="coerce"),
    })
    mapping = [
        ("sales_target", "Sales Target"),
        ("sales_achieved", "Sales Achieved"),
        ("nob_target", "NOB Target"),
        ("nob_achieved", "NOB Achieved"),
        ("abv_target", "ABV Target"),
        ("abv_actual", "ABV Actual"),
        ("gp_target", "GP Target"),
        ("gp_actual", "GP Actual"),
    ]
    for key, label in mapping:
        if found[key]:
            df[label] = pd.to_numeric(raw[found[key]], errors="coerce")
        else:
            df[label] = np.nan
    # Keep dated rows that contain at least one sales value.
    df = df[df["Date"].notna() & (df["Sales Achieved"].notna() | df["Sales Target"].notna())].copy()
    return df


def _load_web_data(input_file):
    if not input_file:
        return pd.DataFrame()
    book = pd.ExcelFile(input_file, engine="openpyxl")
    records = []
    controlled_sheet = next((s for s in ["Data Update", "Data"] if s in book.sheet_names), None)
    if controlled_sheet:
        raw = pd.read_excel(input_file, sheet_name=controlled_sheet, engine="openpyxl")
        frame = _standardize_web_source(raw)
        if frame is not None:
            records.append(frame)
    else:
        branch_sheets = [s for s in book.sheet_names if str(s).strip().isdigit()]
        for sheet in branch_sheets:
            raw = pd.read_excel(input_file, sheet_name=sheet, engine="openpyxl")
            frame = _standardize_web_source(raw, fixed_branch=sheet)
            if frame is not None:
                frame["Branch"] = frame["Branch"].map(
                    lambda code: f"{code} - {BRANCH_NAMES.get(str(code), 'NAME NOT MAPPED')}"
                )
                records.append(frame)
    if not records:
        return pd.DataFrame()
    data = pd.concat(records, ignore_index=True)
    data["Branch"] = data["Branch"].astype(str).str.strip()
    data = data[~data["Branch"].isin(["", "nan", "None"])]
    data = data.sort_values(["Date", "Branch"]).reset_index(drop=True)
    return data


def _money(v):
    return f"SAR {float(v):,.0f}"


def _pct(v):
    return f"{float(v):.1%}"


def _empty_figure(title, message="Upload an Excel file to populate this chart."):
    fig = go.Figure()
    fig.add_annotation(text=message, x=0.5, y=0.5, xref="paper", yref="paper", showarrow=False,
                       font=dict(size=14, color="#64748B"))
    fig.update_layout(title=title, template="plotly_white", height=330,
                      margin=dict(l=35, r=20, t=55, b=35),
                      paper_bgcolor="white", plot_bgcolor="white")
    return fig


def _card(label, value, tone="navy", foot=""):
    return (
        f'<div class="kpi-card {tone}">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value}</div>'
        f'<div class="kpi-foot">{foot or "&nbsp;"}</div>'
        f'</div>'
    )


def _kpi_html(values):
    cards = [
        _card("SALES TARGET", _money(values["sales_target"]), "blue"),
        _card("SALES ACHIEVED", _money(values["sales_actual"]), "green"),
        _card("SALES ACHIEVEMENT", _pct(values["sales_pct"]), "orange"),
        _card("SALES VARIANCE", _money(values["variance"]), "green" if values["variance"] >= 0 else "red"),
        _card("NOB TARGET", f'{values["nob_target"]:,.0f}', "purple"),
        _card("NOB ACHIEVED", f'{values["nob_actual"]:,.0f}', "purple"),
        _card("NOB ACHIEVEMENT", _pct(values["nob_pct"]), "orange"),
        _card("ABV TARGET", _money(values["abv_target"]), "blue"),
        _card("ABV ACTUAL", _money(values["abv_actual"]), "green"),
        _card("PREVIOUS PERIOD SALES", _money(values["previous_sales"]), "navy"),
        _card("PERIOD CHANGE", _pct(values["period_change"]), "green" if values["period_change"] >= 0 else "red"),
        _card("TOP BRANCH", values["top_branch"] or "Not available", "navy"),
    ]
    return '<div class="kpi-grid">' + "".join(cards) + "</div>"


def _insights_html(frame):
    if frame.empty:
        return '<div class="empty-panel">No data for the selected filters.</div>'
    grouped = frame.groupby("Branch", as_index=False).agg(
        Target=("Sales Target", "sum"),
        Actual=("Sales Achieved", "sum"),
        NOB_Target=("NOB Target", "sum"),
        NOB_Actual=("NOB Achieved", "sum"),
    )
    grouped["Pct"] = np.where(grouped["Target"] != 0, grouped["Actual"] / grouped["Target"], 0)
    grouped["NOB_Pct"] = np.where(grouped["NOB_Target"] != 0, grouped["NOB_Actual"] / grouped["NOB_Target"], 0)
    rows = []
    for _, r in grouped.sort_values("Pct").head(6).iterrows():
        pct = float(r["Pct"])
        gap = float(r["Actual"] - r["Target"])
        if pct >= 1:
            tag, cls, action = "▲ POSITIVE", "positive", "Maintain the current pace."
        elif pct >= 0.85:
            tag, cls, action = "● WATCH", "watch", "Close the remaining sales gap and review ABV."
        else:
            tag, cls, action = "⚠ CRITICAL", "critical", "Start a daily recovery plan and review conversion."
        rows.append(
            f'<div class="insight-row"><span class="insight-tag {cls}">{tag}</span>'
            f'<span><b>{r["Branch"]}</b> — Sales {_pct(pct)}; gap {_money(abs(gap))}; '
            f'bills {_pct(float(r["NOB_Pct"]))}. {action}</span></div>'
        )
    return '<div class="insights">' + "".join(rows) + "</div>"


def _forecast_html(values):
    items = [
        ("Elapsed Days", f'{values["elapsed_days"]} / {values["days_in_month"]}'),
        ("Days Remaining", f'{values["days_remaining"]}'),
        ("Month-End Forecast", _money(values["month_end_forecast"])),
        ("Projected Target", _money(values["projected_target"])),
        ("Forecast Gap", _money(values["forecast_gap"])),
        ("Remaining Target", _money(values["remaining_target"])),
        ("Required Daily Sales", _money(values["required_daily_sales"])),
    ]
    html = ''.join(f'<div class="forecast-card"><span>{a}</span><b>{b}</b></div>' for a, b in items)
    return '<div class="forecast-grid">' + html + '</div>'


def _heatmap_html(frame):
    if frame.empty:
        return '<div class="empty-panel">No branch performance data.</div>'
    g = frame.groupby("Branch", as_index=False).agg(
        Target=("Sales Target", "sum"), Actual=("Sales Achieved", "sum"),
        NOB_Target=("NOB Target", "sum"), NOB_Actual=("NOB Achieved", "sum"),
    )
    g["SalesPct"] = np.where(g["Target"] != 0, g["Actual"] / g["Target"], 0)
    g["Gap"] = g["Actual"] - g["Target"]
    g["NOBPct"] = np.where(g["NOB_Target"] != 0, g["NOB_Actual"] / g["NOB_Target"], 0)
    trs = []
    for _, r in g.sort_values("SalesPct", ascending=False).iterrows():
        pct = float(r["SalesPct"])
        status = "ON TRACK" if pct >= 1 else ("WATCH" if pct >= 0.85 else "CRITICAL")
        cls = "positive" if pct >= 1 else ("watch" if pct >= 0.85 else "critical")
        gap_cls = "positive" if float(r["Gap"]) >= 0 else "critical"
        trs.append(
            f'<tr><td class="branch-cell">{r["Branch"]}</td>'
            f'<td class="heat {cls}">{_pct(pct)}</td>'
            f'<td class="heat {gap_cls}">{_money(float(r["Gap"]))}</td>'
            f'<td>{_pct(float(r["NOBPct"]))}</td>'
            f'<td class="status {cls}">{status}</td></tr>'
        )
    return (
        '<div class="table-wrap"><table class="heatmap-table"><thead><tr>'
        '<th>BRANCH</th><th>SALES %</th><th>GAP • SAR</th><th>NOB %</th><th>STATUS</th>'
        '</tr></thead><tbody>' + ''.join(trs) + '</tbody></table></div>'
    )


def _filtered_frame(data, branch, month, date_filter):
    frame = data.copy()
    if branch and branch != "All Branches":
        frame = frame[frame["Branch"] == branch]
    if month and month != "All Months":
        frame = frame[frame["Date"].dt.strftime("%b-%Y") == month]
    if date_filter and date_filter != "All Dates":
        try:
            dt = pd.to_datetime(date_filter, format="%d-%b-%Y")
        except Exception:
            dt = pd.to_datetime(date_filter, errors="coerce")
        if pd.notna(dt):
            frame = frame[frame["Date"].dt.normalize() == pd.Timestamp(dt).normalize()]
    return frame


def _dashboard_values(data, frame, selected_month):
    def total(col):
        v = frame[col].sum(min_count=1)
        return 0.0 if pd.isna(v) else float(v)

    sales_target = total("Sales Target")
    sales_actual = total("Sales Achieved")
    nob_target = total("NOB Target")
    nob_actual = total("NOB Achieved")
    sales_pct = sales_actual / sales_target if sales_target else 0.0
    nob_pct = nob_actual / nob_target if nob_target else 0.0
    variance = sales_actual - sales_target
    abv_actual = sales_actual / nob_actual if nob_actual else 0.0
    abv_target = float(frame["ABV Target"].dropna().mean()) if frame["ABV Target"].notna().any() else 0.0

    month_period = None
    if selected_month and selected_month != "All Months":
        try:
            month_period = pd.Period(pd.to_datetime(selected_month, format="%b-%Y"), freq="M")
        except Exception:
            month_period = None
    if month_period is None and not frame.empty:
        month_period = frame["Date"].max().to_period("M")

    previous_sales = 0.0
    period_change = 0.0
    elapsed_days = int(frame["Date"].max().day) if not frame.empty else 0
    days_in_month = int(month_period.days_in_month) if month_period is not None else 0

    if month_period is not None:
        prev_period = month_period - 1
        prev = data[data["Date"].dt.to_period("M") == prev_period].copy()
        if not frame.empty and frame["Branch"].nunique() == 1:
            prev = prev[prev["Branch"] == frame["Branch"].iloc[0]]
        if elapsed_days:
            cutoff = min(elapsed_days, prev_period.days_in_month)
            prev = prev[prev["Date"].dt.day <= cutoff]
        p = prev["Sales Achieved"].sum(min_count=1)
        previous_sales = 0.0 if pd.isna(p) else float(p)
        period_change = (sales_actual - previous_sales) / previous_sales if previous_sales else 0.0

    days_remaining = max(days_in_month - elapsed_days, 0)
    month_end_forecast = sales_actual / elapsed_days * days_in_month if elapsed_days else 0.0
    projected_target = sales_target / elapsed_days * days_in_month if elapsed_days else 0.0
    forecast_gap = month_end_forecast - projected_target
    remaining_target = max(projected_target - sales_actual, 0.0)
    required_daily_sales = remaining_target / days_remaining if days_remaining else 0.0

    top_branch = ""
    if not frame.empty:
        b = frame.groupby("Branch", as_index=False).agg(Target=("Sales Target", "sum"), Actual=("Sales Achieved", "sum"))
        b["Pct"] = np.where(b["Target"] != 0, b["Actual"] / b["Target"], 0)
        if not b.empty:
            row = b.sort_values("Pct", ascending=False).iloc[0]
            top_branch = f'{row["Branch"]} • {_pct(float(row["Pct"]))}'

    return dict(
        sales_target=sales_target, sales_actual=sales_actual, sales_pct=sales_pct, variance=variance,
        nob_target=nob_target, nob_actual=nob_actual, nob_pct=nob_pct,
        abv_target=abv_target, abv_actual=abv_actual,
        previous_sales=previous_sales, period_change=period_change, top_branch=top_branch,
        elapsed_days=elapsed_days, days_in_month=days_in_month, days_remaining=days_remaining,
        month_end_forecast=month_end_forecast, projected_target=projected_target,
        forecast_gap=forecast_gap, remaining_target=remaining_target,
        required_daily_sales=required_daily_sales,
    )


def render_live_dashboard(input_file, branch, month, date_filter):
    if not input_file:
        empty = _empty_figure("Daily Sales Trend")
        return (
            '<div class="empty-panel hero-empty">Upload an MTD Excel file to load the live dashboard.</div>',
            '<div class="empty-panel">Management insights will appear here.</div>',
            '<div class="empty-panel">Forecast & pace will appear here.</div>',
            empty,
            _empty_figure("Branch Target vs Achieved"),
            _empty_figure("Current vs Previous Period"),
            '<div class="empty-panel">Branch performance heatmap will appear here.</div>',
            _empty_figure("Bill / NOB Achievement"),
            pd.DataFrame(columns=["Branch", "Date", "Sales Target", "Sales Achieved", "Gap", "Sales %", "Flag"]),
            "Waiting for source file.",
        )
    try:
        data = _load_web_data(input_file)
        if data.empty:
            raise ValueError("No usable daily branch data found in the workbook.")
        frame = _filtered_frame(data, branch, month, date_filter)
        values = _dashboard_values(data, frame, month)

        # Daily Sales Trend
        daily = frame.groupby("Date", as_index=False).agg(Target=("Sales Target", "sum"), Actual=("Sales Achieved", "sum"))
        daily_fig = go.Figure()
        daily_fig.add_trace(go.Scatter(x=daily["Date"], y=daily["Target"], mode="lines+markers", name="Target"))
        daily_fig.add_trace(go.Scatter(x=daily["Date"], y=daily["Actual"], mode="lines+markers", name="Achieved"))
        daily_fig.update_layout(template="plotly_white", height=340, margin=dict(l=35,r=20,t=45,b=35),
                                legend=dict(orientation="h", y=1.12), yaxis_title="SAR", xaxis_title="")

        # Branch Target vs Achieved
        bg = frame.groupby("Branch", as_index=False).agg(Target=("Sales Target", "sum"), Actual=("Sales Achieved", "sum"))
        branch_fig = go.Figure()
        branch_fig.add_trace(go.Bar(x=bg["Branch"], y=bg["Target"], name="Target"))
        branch_fig.add_trace(go.Bar(x=bg["Branch"], y=bg["Actual"], name="Achieved"))
        branch_fig.update_layout(barmode="group", template="plotly_white", height=350,
                                 margin=dict(l=35,r=20,t=45,b=90), legend=dict(orientation="h", y=1.12),
                                 yaxis_title="SAR", xaxis_tickangle=-25)

        # Current vs previous period by branch.
        current_period = None
        if month and month != "All Months":
            try:
                current_period = pd.Period(pd.to_datetime(month, format="%b-%Y"), freq="M")
            except Exception:
                pass
        if current_period is None:
            current_period = data["Date"].max().to_period("M")
        prev_period = current_period - 1
        current_cmp = data[data["Date"].dt.to_period("M") == current_period].copy()
        prev_cmp = data[data["Date"].dt.to_period("M") == prev_period].copy()
        if branch and branch != "All Branches":
            current_cmp = current_cmp[current_cmp["Branch"] == branch]
            prev_cmp = prev_cmp[prev_cmp["Branch"] == branch]
        elapsed = int(current_cmp["Date"].max().day) if not current_cmp.empty else current_period.days_in_month
        prev_cmp = prev_cmp[prev_cmp["Date"].dt.day <= min(elapsed, prev_period.days_in_month)]
        c = current_cmp.groupby("Branch", as_index=False)["Sales Achieved"].sum().rename(columns={"Sales Achieved":"Current"})
        p = prev_cmp.groupby("Branch", as_index=False)["Sales Achieved"].sum().rename(columns={"Sales Achieved":"Previous"})
        comp = pd.merge(c, p, on="Branch", how="outer").fillna(0)
        comp_fig = go.Figure()
        comp_fig.add_trace(go.Bar(x=comp["Branch"], y=comp["Previous"], name=f"{prev_period.to_timestamp():%b-%Y}"))
        comp_fig.add_trace(go.Bar(x=comp["Branch"], y=comp["Current"], name=f"{current_period.to_timestamp():%b-%Y}"))
        comp_fig.update_layout(barmode="group", template="plotly_white", height=350,
                               margin=dict(l=35,r=20,t=45,b=90), legend=dict(orientation="h", y=1.12),
                               yaxis_title="SAR", xaxis_tickangle=-25)

        # NOB achievement by branch.
        ng = frame.groupby("Branch", as_index=False).agg(Target=("NOB Target", "sum"), Actual=("NOB Achieved", "sum"))
        ng["Pct"] = np.where(ng["Target"] != 0, ng["Actual"] / ng["Target"], 0)
        nob_fig = go.Figure(go.Bar(x=ng["Pct"], y=ng["Branch"], orientation="h", text=[_pct(x) for x in ng["Pct"]], textposition="outside"))
        nob_fig.update_layout(template="plotly_white", height=max(330, 55 * max(len(ng), 1)),
                              margin=dict(l=35,r=45,t=45,b=35), xaxis_tickformat=".0%", xaxis_title="Achievement",
                              yaxis_title="")

        # Daily exception report.
        ex = frame.groupby(["Branch", "Date"], as_index=False).agg(
            **{"Sales Target": ("Sales Target", "sum"), "Sales Achieved": ("Sales Achieved", "sum")}
        )
        ex["Gap"] = ex["Sales Achieved"] - ex["Sales Target"]
        ex["Sales %"] = np.where(ex["Sales Target"] != 0, ex["Sales Achieved"] / ex["Sales Target"], 0)
        ex["Flag"] = np.where(ex["Sales %"] >= 1, "ON TRACK", np.where(ex["Sales %"] >= .85, "WATCH", "CRITICAL"))
        ex = ex[ex["Flag"] != "ON TRACK"].sort_values(["Sales %", "Date"]).head(40)
        ex["Date"] = ex["Date"].dt.strftime("%d-%b-%Y")
        ex["Sales %"] = ex["Sales %"].map(lambda x: f"{x:.1%}")
        for col in ["Sales Target", "Sales Achieved", "Gap"]:
            ex[col] = ex[col].map(lambda x: round(float(x), 2))

        status = f"Live dashboard loaded: {len(data):,} source rows | showing {len(frame):,} filtered rows."
        return (
            _kpi_html(values), _insights_html(frame), _forecast_html(values), daily_fig, branch_fig,
            comp_fig, _heatmap_html(frame), nob_fig, ex, status
        )
    except Exception as exc:
        msg = f"Could not build live dashboard: {exc}"
        return (
            f'<div class="empty-panel critical">{msg}</div>', '<div class="empty-panel">No insights.</div>',
            '<div class="empty-panel">No forecast.</div>', _empty_figure("Daily Sales Trend", msg),
            _empty_figure("Branch Target vs Achieved", msg), _empty_figure("Current vs Previous Period", msg),
            '<div class="empty-panel">No heatmap.</div>', _empty_figure("Bill / NOB Achievement", msg),
            pd.DataFrame(), msg
        )


def initialize_live_dashboard(input_file):
    if not input_file:
        return (
            gr.Dropdown(choices=["All Branches"], value="All Branches"),
            gr.Dropdown(choices=["All Months"], value="All Months"),
            gr.Dropdown(choices=["All Dates"], value="All Dates"),
            *render_live_dashboard(None, "All Branches", "All Months", "All Dates")
        )
    try:
        data = _load_web_data(input_file)
        if data.empty:
            raise ValueError("No usable data found.")
        branches = ["All Branches"] + sorted(data["Branch"].dropna().unique().tolist())
        periods = sorted(data["Date"].dt.to_period("M").dropna().unique().tolist())
        months = [p.to_timestamp().strftime("%b-%Y") for p in periods]
        default_month = months[-1] if months else "All Months"
        month_choices = ["All Months"] + months
        dates = sorted(data[data["Date"].dt.strftime("%b-%Y") == default_month]["Date"].dropna().unique().tolist())
        date_choices = ["All Dates"] + [pd.Timestamp(d).strftime("%d-%b-%Y") for d in dates]
        dashboard_outputs = render_live_dashboard(input_file, "All Branches", default_month, "All Dates")
        return (
            gr.Dropdown(choices=branches, value="All Branches"),
            gr.Dropdown(choices=month_choices, value=default_month),
            gr.Dropdown(choices=date_choices, value="All Dates"),
            *dashboard_outputs
        )
    except Exception as exc:
        return (
            gr.Dropdown(choices=["All Branches"], value="All Branches"),
            gr.Dropdown(choices=["All Months"], value="All Months"),
            gr.Dropdown(choices=["All Dates"], value="All Dates"),
            *render_live_dashboard(None, "All Branches", "All Months", "All Dates")[:-1],
            f"Could not initialize dashboard: {exc}",
        )


def update_date_choices(input_file, month):
    if not input_file:
        return gr.Dropdown(choices=["All Dates"], value="All Dates")
    try:
        data = _load_web_data(input_file)
        if month and month != "All Months":
            data = data[data["Date"].dt.strftime("%b-%Y") == month]
        dates = sorted(data["Date"].dropna().unique().tolist())
        choices = ["All Dates"] + [pd.Timestamp(d).strftime("%d-%b-%Y") for d in dates]
        return gr.Dropdown(choices=choices, value="All Dates")
    except Exception:
        return gr.Dropdown(choices=["All Dates"], value="All Dates")


def inspect_upload_secure(uploaded_file, request: gr.Request):
    role, _ = _role_from_request(request)
    if not uploaded_file:
        return gr.Dropdown(choices=[], value=None), gr.Dropdown(choices=[], value=None), "Upload an Excel source file first."
    try:
        months, branches = _extract_source_info(uploaded_file)
        persisted_month, persisted_status = _read_control_settings(uploaded_file)
        default_month = persisted_month or (months[-2] if len(months) >= 2 else (months[-1] if months else None))
        default_model = branches[0] if branches else None
        if role == "ADMIN":
            message = f"Admin source loaded: {len(branches)} branch(es). You may set the controlled period and branch control."
        else:
            if persisted_month and persisted_status:
                message = f"Controlled workbook loaded: {persisted_month} | {persisted_status}. Admin controls are locked for your account."
            else:
                message = "Dashboard preview is available, but this workbook is not Admin-controlled yet. Ask Admin to initialize the period before generating a new XLSM."
        return gr.Dropdown(choices=months, value=default_month), gr.Dropdown(choices=branches, value=default_model), message
    except Exception as exc:
        return gr.Dropdown(choices=[], value=None), gr.Dropdown(choices=[], value=None), f"Could not read the source file: {exc}"


def build_dashboard_secure(input_file, selected_month, selected_status, add_branch, new_branch, model_branch, request: gr.Request):
    """Enforce Admin-only period/new-branch changes on the server, not just in the UI."""
    role, username = _role_from_request(request)
    if role == "ADMIN":
        return _build_dashboard_core(input_file, selected_month, selected_status, add_branch, new_branch, model_branch)

    persisted_month, persisted_status = _read_control_settings(input_file)
    if not persisted_month or not persisted_status:
        raise gr.Error(
            "Admin setup required. Normal users cannot OPEN/CLOSE a period or initialize a new company/branch. "
            "Please ask Admin to generate the controlled workbook once."
        )
    # Ignore any browser-manipulated Admin values and force persisted workbook control.
    return _build_dashboard_core(input_file, persisted_month, persisted_status, "NO", "", "")


def page_role(request: gr.Request):
    role, username = _role_from_request(request)
    if role == "ADMIN":
        badge = f'<div class="role-badge admin">ADMIN • {username} &nbsp;|&nbsp; Period & Branch controls enabled</div>'
        return badge, gr.update(visible=True)
    badge = f'<div class="role-badge user">USER • {username} &nbsp;|&nbsp; View / filter / download only</div>'
    return badge, gr.update(visible=False)


CSS = r"""
:root { --navy:#153B5B; --blue:#2673C9; --teal:#12A7A0; --green:#24A148; --orange:#F59E0B; --red:#D64545; --bg:#F3F7FB; --muted:#64748B; }
.gradio-container { max-width: 1480px !important; margin:auto !important; background:#F3F7FB !important; }
#app-title h1 { color:#153B5B; margin-bottom:0 !important; font-weight:800; }
#app-subtitle { color:#64748B; margin-top:0 !important; }
.role-badge { padding:10px 16px; border-radius:10px; font-weight:700; margin-bottom:10px; display:inline-block; }
.role-badge.admin { background:#EAF8F3; color:#13795B; border:1px solid #8CD2B5; }
.role-badge.user { background:#EAF3FF; color:#155AA8; border:1px solid #A9CBF3; }
.admin-box { border:1px solid #F4C36A !important; background:#FFFDF7 !important; border-radius:12px !important; padding:10px !important; }
.section-title { background:#2673C9; color:white; font-weight:800; padding:10px 14px; border-radius:9px 9px 0 0; margin-top:10px; letter-spacing:.2px; }
.kpi-grid { display:grid; grid-template-columns:repeat(6,minmax(150px,1fr)); gap:10px; margin:10px 0 16px; }
.kpi-card { background:white; border:1px solid #D7E2EC; border-radius:12px; padding:13px 12px; box-shadow:0 2px 8px rgba(21,59,91,.06); min-height:108px; }
.kpi-card.blue { border-top:4px solid #2673C9; } .kpi-card.green { border-top:4px solid #24A148; }
.kpi-card.orange { border-top:4px solid #F59E0B; } .kpi-card.red { border-top:4px solid #D64545; }
.kpi-card.purple { border-top:4px solid #6D4CC2; } .kpi-card.navy { border-top:4px solid #153B5B; }
.kpi-label { color:#64748B; font-size:12px; font-weight:800; } .kpi-value { color:#153B5B; font-size:22px; font-weight:850; margin-top:9px; }
.kpi-foot { color:#94A3B8; font-size:10px; margin-top:5px; }
.insights { background:white; border:1px solid #D7E2EC; border-radius:0 0 10px 10px; padding:8px 12px; }
.insight-row { display:flex; gap:10px; align-items:center; border-bottom:1px solid #EEF2F7; padding:9px 0; color:#203040; }
.insight-row:last-child { border-bottom:none; } .insight-tag { min-width:94px; padding:5px 8px; border-radius:7px; text-align:center; font-size:11px; font-weight:800; }
.positive { background:#DDF4E9 !important; color:#116149 !important; } .watch { background:#FFF1CC !important; color:#8A5300 !important; } .critical { background:#FDE2E2 !important; color:#B42318 !important; }
.forecast-grid { display:grid; grid-template-columns:repeat(7,minmax(120px,1fr)); gap:8px; background:white; border:1px solid #D7E2EC; padding:10px; border-radius:0 0 10px 10px; }
.forecast-card { padding:9px; background:#F8FAFC; border-radius:8px; border:1px solid #E2E8F0; display:flex; flex-direction:column; gap:4px; }
.forecast-card span { color:#64748B; font-size:11px; font-weight:700; } .forecast-card b { color:#153B5B; font-size:16px; }
.empty-panel { background:white; border:1px dashed #CBD5E1; border-radius:10px; padding:22px; color:#64748B; text-align:center; }
.hero-empty { padding:48px 20px; font-size:16px; }
.table-wrap { overflow:auto; background:white; border:1px solid #D7E2EC; border-radius:0 0 10px 10px; }
.heatmap-table { width:100%; border-collapse:collapse; font-size:13px; } .heatmap-table th { background:#153B5B; color:white; padding:10px; text-align:center; }
.heatmap-table td { border-bottom:1px solid #E5EDF4; padding:10px; text-align:center; } .heatmap-table .branch-cell { text-align:left; font-weight:700; color:#153B5B; }
.heatmap-table .heat, .heatmap-table .status { font-weight:800; border-radius:0; }
#generate-btn { background:#F57C20 !important; color:white !important; font-weight:800 !important; min-height:44px; }
footer { display:none !important; }
@media (max-width:1100px){ .kpi-grid{grid-template-columns:repeat(3,1fr)} .forecast-grid{grid-template-columns:repeat(3,1fr)} }
@media (max-width:700px){ .kpi-grid{grid-template-columns:repeat(2,1fr)} .forecast-grid{grid-template-columns:repeat(2,1fr)} }
"""

with gr.Blocks(title="MTD Performance Dashboard") as demo:
    gr.Markdown("# MTD Performance Dashboard", elem_id="app-title")
    gr.Markdown(
        f"**Build: {BUILD_VERSION}** &nbsp; • &nbsp; Live browser dashboard + protected XLSM download",
        elem_id="app-subtitle",
    )
    role_badge = gr.HTML()

    with gr.Row(equal_height=True):
        source_file = gr.File(label="Upload MTD Excel", file_types=[".xlsx", ".xlsm"], type="filepath", scale=2)
        with gr.Column(scale=2):
            source_status = gr.Textbox(label="Source / Control Status", lines=3, interactive=False)
            with gr.Row():
                web_branch = gr.Dropdown(label="Dashboard Branch", choices=["All Branches"], value="All Branches")
                web_month = gr.Dropdown(label="Dashboard Month", choices=["All Months"], value="All Months")
                web_date = gr.Dropdown(label="Dashboard Date", choices=["All Dates"], value="All Dates")

    with gr.Column(visible=False, elem_classes="admin-box") as admin_panel:
        gr.Markdown("### 🔐 Admin Controls — Period / New Company-Branch Setup")
        with gr.Row():
            controlled_month = gr.Dropdown(label="Controlled Month", choices=[])
            period_status = gr.Radio(["OPEN", "CLOSED"], value="OPEN", label="Period Status")
            add_branch = gr.Radio(["NO", "YES"], value="NO", label="Add a new approved branch?")
        with gr.Row():
            new_branch = gr.Textbox(label="New Branch", placeholder="Example: 115 - NEW BRANCH")
            model_branch = gr.Dropdown(label="Copy Targets From Existing Branch", choices=[])

    gr.HTML('<div class="section-title">EXECUTIVE KPI VIEW</div>')
    kpi_panel = gr.HTML('<div class="empty-panel hero-empty">Upload an MTD Excel file to load the live dashboard.</div>')

    gr.HTML('<div class="section-title">MANAGEMENT INSIGHTS</div>')
    insights_panel = gr.HTML('<div class="empty-panel">Management insights will appear here.</div>')

    gr.HTML('<div class="section-title">FORECAST & PACE</div>')
    forecast_panel = gr.HTML('<div class="empty-panel">Forecast & pace will appear here.</div>')

    with gr.Row():
        with gr.Column():
            gr.HTML('<div class="section-title">DAILY SALES TREND</div>')
            daily_plot = gr.Plot(value=_empty_figure("Daily Sales Trend"), show_label=False)
        with gr.Column():
            gr.HTML('<div class="section-title">BRANCH TARGET VS ACHIEVED</div>')
            branch_plot = gr.Plot(value=_empty_figure("Branch Target vs Achieved"), show_label=False)

    with gr.Row():
        with gr.Column():
            gr.HTML('<div class="section-title">CURRENT VS PREVIOUS PERIOD</div>')
            comparison_plot = gr.Plot(value=_empty_figure("Current vs Previous Period"), show_label=False)
        with gr.Column():
            gr.HTML('<div class="section-title">BILL / NOB ACHIEVEMENT</div>')
            nob_plot = gr.Plot(value=_empty_figure("Bill / NOB Achievement"), show_label=False)

    gr.HTML('<div class="section-title">BRANCH PERFORMANCE HEATMAP</div>')
    heatmap_panel = gr.HTML('<div class="empty-panel">Branch performance heatmap will appear here.</div>')

    gr.HTML('<div class="section-title">DAILY EXCEPTION REPORT</div>')
    exceptions_table = gr.Dataframe(
        headers=["Branch", "Date", "Sales Target", "Sales Achieved", "Gap", "Sales %", "Flag"],
        interactive=False, wrap=True, max_height=420,
    )
    dashboard_status = gr.Textbox(label="Live Dashboard Status", interactive=False)

    with gr.Row():
        generate_btn = gr.Button("Generate / Refresh XLSM Download", variant="primary", elem_id="generate-btn", scale=2)
        output_file = gr.File(label="Download Generated Dashboard", scale=2)
    generation_status = gr.Textbox(label="Generation Status", lines=5, interactive=False)

    dashboard_outputs = [
        kpi_panel, insights_panel, forecast_panel, daily_plot, branch_plot,
        comparison_plot, heatmap_panel, nob_plot, exceptions_table, dashboard_status
    ]

    demo.load(fn=page_role, inputs=None, outputs=[role_badge, admin_panel])

    source_file.change(
        fn=inspect_upload_secure,
        inputs=[source_file],
        outputs=[controlled_month, model_branch, source_status],
    )
    source_file.change(
        fn=initialize_live_dashboard,
        inputs=[source_file],
        outputs=[web_branch, web_month, web_date] + dashboard_outputs,
    )

    web_month.change(fn=update_date_choices, inputs=[source_file, web_month], outputs=[web_date])
    for control in [web_branch, web_month, web_date]:
        control.change(
            fn=render_live_dashboard,
            inputs=[source_file, web_branch, web_month, web_date],
            outputs=dashboard_outputs,
        )

    generate_btn.click(
        fn=build_dashboard_secure,
        inputs=[source_file, controlled_month, period_status, add_branch, new_branch, model_branch],
        outputs=[output_file, generation_status],
    )


if __name__ == "__main__":
    _validate_credentials_config()
    port = int(os.getenv("PORT", "7860"))
    demo.queue().launch(
        server_name="0.0.0.0",
        server_port=port,
        auth=_auth_user,
        auth_message="Sign in with your Admin or User account.",
        css=CSS,
        show_error=True,
    )
